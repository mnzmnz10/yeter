from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form, Depends, Request, Cookie
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from decimal import Decimal, InvalidOperation
from bson import ObjectId
import os
import uuid
import pandas as pd
import requests
import logging
from io import BytesIO
import hashlib
import secrets

from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image as PILImage
from pathlib import Path
from dotenv import load_dotenv
import io
import openpyxl
from openpyxl.styles import PatternFill

# Load environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create database indexes for better performance
async def create_indexes():
    """Create database indexes for optimal performance with large datasets"""
    try:
        # Products collection indexes
        await db.products.create_index("name")  # For name search
        await db.products.create_index("company_id")  # For company filtering
        await db.products.create_index("category_id")  # For category filtering 
        await db.products.create_index([("name", "text"), ("description", "text")])  # For text search
        
        # Companies collection indexes
        await db.companies.create_index("name")
        
        # Categories collection indexes
        await db.categories.create_index("name")
        
        # Quotes collection indexes
        await db.quotes.create_index("customer_name")
        await db.quotes.create_index("created_at")
        
        logger.info("Database indexes created successfully")
    except Exception as e:
        logger.error(f"Error creating indexes: {e}")

# Create the main app
app = FastAPI(title="Karavan Elektrik Ekipmanları Fiyat Karşılaştırma API")

# Initialize Sarf Malzemeleri Category
async def create_supplies_category():
    """Create default 'Sarf Malzemeleri' category if it doesn't exist"""
    try:
        # Check if supplies category already exists
        existing_category = await db.categories.find_one({"name": "Sarf Malzemeleri"})
        if not existing_category:
            supplies_category = {
                "id": "sarf-malzemeleri-category",
                "name": "Sarf Malzemeleri",
                "description": "Üretimde kullanılan sarf malzemeleri (tutkal, vida, kablo vb.)",
                "color": "#f97316",  # Orange color
                "created_at": datetime.now(timezone.utc),
                "is_deletable": False  # Prevent deletion
            }
            
            await db.categories.insert_one(supplies_category)
            logger.info("Sarf Malzemeleri category created successfully")
            return True
        else:
            # Update existing category to make it non-deletable
            await db.categories.update_one(
                {"name": "Sarf Malzemeleri"},
                {"$set": {"is_deletable": False, "color": "#f97316"}}
            )
            logger.info("Sarf Malzemeleri category updated to non-deletable")
            return True
    except Exception as e:
        logger.error(f"Error creating supplies category: {e}")
        return False

@app.on_event("startup")
async def startup_event():
    """Initialize database indexes and create default categories on startup"""
    await create_indexes()
    await create_supplies_category()
    await create_default_admin()
    logger.info("Application startup completed")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add GZip compression for better performance
app.add_middleware(GZipMiddleware, minimum_size=1000)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Pydantic Models
class Company(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CompanyCreate(BaseModel):
    name: str

class Product(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    company_id: str
    category_id: Optional[str] = None
    brand: str = ""  # Marka alanı
    description: Optional[str] = None
    image_url: Optional[str] = None
    list_price: Decimal
    discounted_price: Optional[Decimal] = None
    currency: str
    list_price_try: Optional[Decimal] = None
    discounted_price_try: Optional[Decimal] = None
    is_favorite: bool = False
    stock_quantity: Optional[int] = None  # Sadece favori ürünler için stok takibi
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    name: str
    company_id: str
    category_id: Optional[str] = None
    brand: str = ""  # Yeni marka alanı
    description: Optional[str] = None
    image_url: Optional[str] = None
    list_price: Decimal
    discounted_price: Optional[Decimal] = None
    currency: str
    is_favorite: bool = False

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    color: Optional[str] = None
    sort_order: int = 0  # Kategori sıralama numarası
    is_deletable: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: Optional[str] = None
    sort_order: Optional[int] = 0  # Yeni kategori için varsayılan sıra

class CategoryGroup(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    color: Optional[str] = "#6B7280"  # Default gray color
    category_ids: List[str] = []  # Bu gruba dahil kategoriler
    sort_order: int = 0  # Kategori grup sıralama numarası
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryGroupCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: Optional[str] = "#6B7280"
    category_ids: List[str] = []
    sort_order: Optional[int] = 0  # Yeni kategori grubu için varsayılan sıra

class CategoryGroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None
    category_ids: Optional[List[str]] = None
    sort_order: Optional[int] = None  # Sıralama güncellemesi için

# Upload History Models
class UploadHistory(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    company_name: str
    filename: str
    upload_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    total_products: int
    new_products: int
    updated_products: int
    currency_distribution: Dict[str, int]  # Currency -> count
    price_changes: List[Dict[str, Any]] = []  # Price change details
    status: str = "completed"  # completed, failed, processing

# Package Models
class Package(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    sale_price: Optional[Decimal] = None
    discount_percentage: float = 0  # Paket indirimi
    labor_cost: float = 0  # İşçilik maliyeti
    image_url: Optional[str] = None
    is_pinned: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PackageCreate(BaseModel):
    name: str
    description: Optional[str] = None
    sale_price: Optional[Decimal] = None
    discount_percentage: float = 0  # Paket indirimi
    labor_cost: float = 0  # İşçilik maliyeti
    image_url: Optional[str] = None
    is_pinned: bool = False

class PackageProduct(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    package_id: str
    product_id: str
    quantity: int = 1
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PackageProductCreate(BaseModel):
    product_id: str
    quantity: int = 1

class PackageWithProducts(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    sale_price: Optional[Decimal] = None
    discount_percentage: float = 0  # Paket indirimi
    labor_cost: float = 0  # İşçilik maliyeti
    image_url: Optional[str] = None
    created_at: datetime
    products: List[Dict[str, Any]] = []
    supplies: List[Dict[str, Any]] = []  # Sarf malzemeleri
    total_discounted_price: Optional[Decimal] = None
    total_discounted_price_with_supplies: Optional[Decimal] = None
    status: str
class UploadHistoryResponse(BaseModel):
    id: str
    company_id: str
    company_name: str
    filename: str
    upload_date: datetime
    total_products: int
    new_products: int
    updated_products: int
    currency_distribution: Dict[str, int]
    price_changes: List[Dict[str, Any]]
    status: str

class QuoteCreate(BaseModel):
    name: str
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    discount_percentage: float = 0
    labor_cost: float = 0  # İşçilik maliyeti
    products: List[Dict[str, Any]]  # Product objects with ID and quantity
    notes: Optional[str] = None
class PackageSupply(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    package_id: str
    product_id: str
    quantity: int = 1
    note: Optional[str] = None  # Sarf malzemesi notu
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PackageSupplyCreate(BaseModel):
    product_id: str
    quantity: int = 1
    note: Optional[str] = None

class QuoteResponse(BaseModel):
    id: str
    name: str
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    discount_percentage: float
    labor_cost: float = 0  # İşçilik maliyeti
    total_list_price: float
    total_discounted_price: float 
    total_net_price: float
    products: List[Dict[str, Any]]
    notes: Optional[str] = None
    created_at: str
    status: str = "active"

class ExchangeRate(BaseModel):
    currency: str
    rate_to_try: Decimal
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Authentication Models
class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    success: bool
    message: str
    session_token: Optional[str] = None

class User(BaseModel):
    id: str
    username: str
    password_hash: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = True

# Currency conversion service
class CurrencyService:
    def __init__(self):
        self.api_url = "https://api.exchangerate-api.com/v4/latest/TRY"
        self.rates_cache = {}
        self.last_update = None
    
    async def get_exchange_rates(self) -> Dict[str, Decimal]:
        """Get current exchange rates from API"""
        try:
            response = requests.get(self.api_url, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            # Convert to rates from other currencies to TRY
            base_rates = data.get('rates', {})
            try_to_usd = base_rates.get('USD', 1)
            try_to_eur = base_rates.get('EUR', 1)
            
            # Calculate USD and EUR to TRY
            usd_to_try = Decimal('1') / Decimal(str(try_to_usd)) if try_to_usd else Decimal('1')
            eur_to_try = Decimal('1') / Decimal(str(try_to_eur)) if try_to_eur else Decimal('1')
            
            rates = {
                'USD': usd_to_try,
                'EUR': eur_to_try,
                'TRY': Decimal('1'),
                'GBP': Decimal('1') / Decimal(str(base_rates.get('GBP', 1))) if base_rates.get('GBP') else Decimal('1')
            }
            
            self.rates_cache = rates
            self.last_update = datetime.now(timezone.utc)
            
            # Save to database
            for currency, rate in rates.items():
                await db.exchange_rates.replace_one(
                    {'currency': currency},
                    {
                        'currency': currency,
                        'rate_to_try': float(rate),
                        'updated_at': self.last_update
                    },
                    upsert=True
                )
            
            logger.info(f"Exchange rates updated: {rates}")
            return rates
            
        except Exception as e:
            logger.error(f"Failed to fetch exchange rates: {e}")
            # Try to get from database as fallback
            rates_from_db = await db.exchange_rates.find().to_list(None)
            if rates_from_db:
                return {rate['currency']: Decimal(str(rate['rate_to_try'])) for rate in rates_from_db}
            
            # Default rates as fallback
            return {
                'USD': Decimal('27.5'),
                'EUR': Decimal('30.0'),
                'TRY': Decimal('1'),
                'GBP': Decimal('35.0')
            }
    
    async def convert_to_try(self, amount: Decimal, from_currency: str) -> Decimal:
        """Convert amount to Turkish Lira"""
        if from_currency.upper() == 'TRY':
            return amount
            
        rates = await self.get_exchange_rates()
        rate = rates.get(from_currency.upper(), Decimal('1'))
        return amount * rate

    async def convert_from_try(self, amount_try: Decimal, to_currency: str) -> Decimal:
        """Convert amount from Turkish Lira to target currency"""
        if to_currency.upper() == 'TRY':
            return amount_try
            
        rates = await self.get_exchange_rates()
        rate = rates.get(to_currency.upper(), Decimal('1'))
        
        # Since rates are TRY to other currency, we need to divide
        if rate > 0:
            return amount_try / rate
        else:
            return amount_try

currency_service = CurrencyService()

# Authentication Service
class AuthService:
    def __init__(self):
        self.sessions = {}  # In-memory session storage - basit cookie based oturum
        
    def hash_password(self, password: str) -> str:
        """Hash password using SHA-256"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def verify_password(self, password: str, password_hash: str) -> bool:
        """Verify password against hash"""
        return self.hash_password(password) == password_hash
    
    def create_session(self, username: str) -> str:
        """Create session token"""
        session_token = secrets.token_urlsafe(32)
        self.sessions[session_token] = {
            'username': username,
            'created_at': datetime.now(timezone.utc),
            'expires_at': datetime.now(timezone.utc) + timedelta(hours=24)  # 24 saat oturum
        }
        return session_token
    
    def validate_session(self, session_token: str) -> Optional[str]:
        """Validate session token and return username if valid"""
        if not session_token or session_token not in self.sessions:
            return None
            
        session = self.sessions[session_token]
        if datetime.now(timezone.utc) > session['expires_at']:
            # Session expired
            del self.sessions[session_token]
            return None
            
        return session['username']
    
    def logout(self, session_token: str) -> bool:
        """Logout user by removing session"""
        if session_token in self.sessions:
            del self.sessions[session_token]
            return True
        return False

auth_service = AuthService()

# Create default admin user
async def create_default_admin():
    """Create default admin user if not exists"""
    try:
        admin_user = await db.users.find_one({"username": "karavan_admin"})
        if not admin_user:
            admin_user = {
                "id": str(uuid.uuid4()),
                "username": "karavan_admin",
                "password_hash": auth_service.hash_password("corlukaravan.5959"),
                "created_at": datetime.now(timezone.utc),
                "is_active": True
            }
            await db.users.insert_one(admin_user)
            logger.info("Default admin user created successfully")
        else:
            logger.info("Default admin user already exists")
    except Exception as e:
        logger.error(f"Error creating default admin user: {e}")

# Authentication dependency
async def get_current_user(session_token: Optional[str] = Cookie(None)):
    """Get current authenticated user"""
    if not session_token:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    username = auth_service.validate_session(session_token)
    if not username:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    
    return username

# Optional authentication dependency (doesn't require auth but returns user if authenticated)
async def get_current_user_optional(session_token: Optional[str] = Cookie(None)):
    """Get current authenticated user (optional)"""
    if not session_token:
        return None
    
    username = auth_service.validate_session(session_token)
    return username

# Database helper function
async def get_db():
    """Get database connection"""
    return db

# Color-based Excel parsing service
class ColorBasedExcelService:
    @staticmethod
    def detect_color_category(fill):
        """Detect color category from cell fill"""
        if not fill or not hasattr(fill, 'start_color'):
            return 'NONE'
            
        color = fill.start_color
        
        # RGB renk kontrolü
        if hasattr(color, 'rgb') and color.rgb:
            try:
                rgb = str(color.rgb).upper()
                
                # Debug için
                logger.debug(f"Checking RGB: {rgb}")
                
                # Kırmızı tonları (Ürün Adı) - FFFF0000 formatını kontrol et
                if 'FFFF0000' in rgb or 'FF0000' in rgb or 'CC0000' in rgb:
                    return 'RED'
                # Mavi tonları (Ürün Açıklaması) - özel mavi tonları
                elif 'FF0070C0' in rgb or '0070C0' in rgb or '0000FF' in rgb or '4472C4' in rgb:
                    return 'BLUE'  
                # Turuncu tonları (İndirimli Fiyat) - Önce daha spesifik kontroller
                elif ('FFFFC000' in rgb or 'FFF4B183' in rgb or 'F4B183' in rgb or 'FF7F00' in rgb or 'FFA500' in rgb or 
                      'FF8C00' in rgb or 'FFFF9900' in rgb or 'FF9900' in rgb):
                    return 'ORANGE'
                # Sarı tonları (Firma) - FFFFFF00 formatını kontrol et
                elif 'FFFFFF00' in rgb or 'FFFF00' in rgb or 'FFC000' in rgb:
                    return 'YELLOW'
                # Yeşil tonları (Liste Fiyatı) - FF00B050 formatını kontrol et
                elif 'FF00B050' in rgb or '00B050' in rgb or '00FF00' in rgb or '008000' in rgb:
                    return 'GREEN'
                    
            except Exception as e:
                logger.warning(f"RGB parsing error: {e}")
                
        # Theme color kontrolü (Excel'de theme color kullanıldığında)
        if hasattr(color, 'theme') and color.theme is not None:
            try:
                theme = color.theme
                logger.debug(f"Checking theme color: {theme}")
                
                # Excel theme color mappings
                if theme == 2:  # Theme color 2 genelde koyu kırmızı
                    return 'RED'
                elif theme == 4:  # Theme color 4 genelde mavi
                    return 'BLUE'  
                elif theme == 5:  # Theme color 5 genelde sarı/accent
                    return 'YELLOW'
                elif theme == 6:  # Theme color 6 genelde yeşil
                    return 'GREEN'
                elif theme == 9:  # Theme color 9 - özel durum: hem yeşil hem turuncu olabilir
                    return 'ORANGE'  # Varsayılan turuncu, ama yeşil de olabilir
                elif theme == 7:  # Theme color 7 genelde turuncu
                    return 'ORANGE'
                    
            except Exception as e:
                logger.warning(f"Theme color parsing error: {e}")
                
        # Index renk kontrolü
        if hasattr(color, 'index') and color.index:
            try:
                index = str(color.index)
                logger.debug(f"Checking color index: {index}")
                # Excel'in standart renk indeksleri
                if index in ['10', '3']:  # Kırmızı
                    return 'RED' 
                elif index in ['12', '5']:  # Mavi
                    return 'BLUE'
                elif index in ['13', '6']:  # Sarı
                    return 'YELLOW'
                elif index in ['11', '4']:  # Yeşil
                    return 'GREEN'
                elif index in ['46', '53']:  # Turuncu
                    return 'ORANGE'
                elif index == '9':  # Bu dosyada index 9 İNDİRİMLİ fiyat için kullanılıyor
                    return 'ORANGE'
            except Exception as e:
                logger.warning(f"Index parsing error: {e}")
                
        return 'NONE'
    
    @staticmethod
    def parse_colored_excel(file_content: bytes, company_name: str = "Unknown") -> List[Dict[str, Any]]:
        """Parse Excel file using color-based column detection"""
        try:
            # Load workbook with data_only=True to get formula results
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            all_products = []
            
            logger.info(f"Processing Excel with {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
            
            for sheet_name in workbook.sheetnames:
                logger.info(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Find header row by looking for colored cells
                header_row = ColorBasedExcelService._find_colored_header_row(sheet)
                if header_row == -1:
                    logger.warning(f"No header found in sheet {sheet_name}, trying to analyze first row as data")
                    # Header yoksa direkt 0. satırı data olarak kabul et ve renkleri analiz et
                    column_mapping = ColorBasedExcelService._analyze_data_row_colors(sheet, 0)
                    if all(val == -1 for val in column_mapping.values()):
                        logger.warning(f"No colored columns found in {sheet_name}, skipping")
                        continue
                    header_row = -1  # Data başlangıcı için -1 kullan
                else:
                    logger.info(f"Found colored header at row {header_row + 1}")
                    # Analyze header colors to map columns
                    column_mapping = ColorBasedExcelService._analyze_header_colors(sheet, header_row)
                
                logger.info(f"Column mapping: {column_mapping}")
                
                # Extract products from this sheet
                sheet_products = ColorBasedExcelService._extract_products_from_sheet(
                    sheet, header_row, column_mapping, company_name
                )
                
                logger.info(f"Extracted {len(sheet_products)} products from {sheet_name}")
                all_products.extend(sheet_products)
            
            logger.info(f"Total products extracted: {len(all_products)}")
            return all_products
            
        except Exception as e:
            logger.error(f"Error in color-based Excel parsing: {e}")
            raise HTTPException(status_code=400, detail=f"Renkli Excel dosyası işlenemedi: {str(e)}")
    
    @staticmethod
    def _find_colored_header_row(sheet) -> int:
        """Find the header row with colored cells"""
        max_search_rows = min(20, sheet.max_row)
        
        for row_idx in range(max_search_rows):
            colored_cells = 0
            non_empty_cells = 0
            meaningful_cells = 0
            
            for col_idx in range(min(10, sheet.max_column)):
                cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                if cell.value and str(cell.value).strip():
                    cell_text = str(cell.value).strip().lower()
                    non_empty_cells += 1
                    
                    # Meaningful header words
                    if any(word in cell_text for word in ['ürün', 'ad', 'açık', 'marka', 'firma', 'fiyat', 'price', 'name']):
                        meaningful_cells += 1
                    
                    color_category = ColorBasedExcelService.detect_color_category(cell.fill)
                    logger.debug(f"Row {row_idx + 1}, Col {col_idx + 1}: '{cell_text}' -> {color_category}")
                    if color_category != 'NONE':
                        colored_cells += 1
            
            logger.debug(f"Row {row_idx + 1}: colored={colored_cells}, meaningful={meaningful_cells}, non_empty={non_empty_cells}")
            
            # Header satırında en az 2 renkli hücre ve 2 anlamlı hücre olmalı
            if colored_cells >= 2 and meaningful_cells >= 2:
                return row_idx
        
        # Fallback: Anlamlı kelimeler içeren satırı bul
        for row_idx in range(max_search_rows):
            meaningful_cells = 0
            for col_idx in range(min(10, sheet.max_column)):
                cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                if cell.value:
                    cell_text = str(cell.value).strip().lower()
                    if any(word in cell_text for word in ['ürün', 'ad', 'açık', 'marka', 'firma', 'fiyat']):
                        meaningful_cells += 1
            
            if meaningful_cells >= 3:
                return row_idx
                
        return -1
    
    @staticmethod
    def detect_currency_from_header(cell_value: str) -> str:
        """Detect currency from header cell text"""
        if not cell_value:
            return 'TRY'
            
        text = str(cell_value).upper().strip()
        
        # Dolar kontrolüne daha fazla varyant ekle
        dollar_keywords = ['$', 'DOLAR', 'DOLLAR', 'USD', 'DOLAR İSARETİ', 'DOLAR IŞARETI', 'AMERİKAN DOLARI', 'AMERIKAN DOLARI']
        if any(keyword in text for keyword in dollar_keywords):
            return 'USD'
        
        # Euro kontrolüne daha fazla varyant ekle  
        euro_keywords = ['€', 'EURO', 'EUR', 'AVRO', 'AVRUPA']
        if any(keyword in text for keyword in euro_keywords):
            return 'EUR'
        
        # TL kontrolüne daha fazla varyant ekle
        tl_keywords = ['₺', 'TL', 'TRY', 'TÜRK', 'LIRA', 'TÜRK LİRASI', 'TURK LIRASI', 'TURKİYE', 'TURKIYE']
        if any(keyword in text for keyword in tl_keywords):
            return 'TRY'
            
        # Varsayılan olarak TRY
        else:
            return 'TRY'

    @staticmethod
    def _analyze_header_colors(sheet, header_row: int) -> Dict[str, int]:
        """Analyze header colors and map to column purposes - ONLY COLOR-BASED"""
        column_mapping = {
            'product_name': -1,
            'description': -1, 
            'brand': -1,  # Yeni: Sarı = Marka
            'company': -1,
            'list_price': -1,
            'discounted_price': -1,
            'currency': 'TRY'  # Varsayılan döviz
        }
        
        for col_idx in range(min(15, sheet.max_column)):
            cell = sheet.cell(row=header_row + 1, column=col_idx + 1)
            if not cell.value:
                continue
                
            color_category = ColorBasedExcelService.detect_color_category(cell.fill)
            
            # SADECE renk kategorilerine göre kolon belirleme - text-based fallback YOK
            if color_category == 'RED':  # Kırmızı = Ürün Adı
                column_mapping['product_name'] = col_idx
            elif color_category == 'BLUE':  # Mavi = Ürün Açıklaması
                column_mapping['description'] = col_idx
            elif color_category == 'YELLOW':  # Sarı = Marka (eskiden Firma)
                column_mapping['brand'] = col_idx
            elif color_category == 'GREEN':  # Yeşil = Liste Fiyatı
                column_mapping['list_price'] = col_idx
                # Yeşil sütun bulunduğunda döviz algıla
                detected_currency = ColorBasedExcelService.detect_currency_from_header(cell.value)
                column_mapping['currency'] = detected_currency
                logger.info(f"Detected currency from header '{cell.value}': {detected_currency}")
            elif color_category == 'ORANGE':  # Turuncu = İndirimli Fiyat
                column_mapping['discounted_price'] = col_idx
            # Diğer renkler veya renksiz veriler ignore edilir
        
        return column_mapping
    
    @staticmethod
    def _analyze_data_row_colors(sheet, data_row: int) -> Dict[str, int]:
        """Analyze first data row colors when no header is found - ONLY COLOR-BASED"""
        column_mapping = {
            'product_name': -1,
            'description': -1, 
            'brand': -1,  # Yeni: Sarı = Marka
            'company': -1,
            'list_price': -1,
            'discounted_price': -1,
            'currency': 'TRY'  # Header yoksa varsayılan TRY
        }
        
        for col_idx in range(min(15, sheet.max_column)):
            cell = sheet.cell(row=data_row + 1, column=col_idx + 1)
            if not cell.value:
                continue
                
            color_category = ColorBasedExcelService.detect_color_category(cell.fill)
            
            # SADECE renk kategorilerine göre kolon belirleme - text-based fallback YOK
            if color_category == 'RED':  # Kırmızı = Ürün Adı
                column_mapping['product_name'] = col_idx
            elif color_category == 'BLUE':  # Mavi = Ürün Açıklaması  
                column_mapping['description'] = col_idx
            elif color_category == 'YELLOW':  # Sarı = Marka (eskiden Firma)
                column_mapping['brand'] = col_idx
            elif color_category == 'GREEN':  # Yeşil = Liste Fiyatı
                column_mapping['list_price'] = col_idx
                # Header yoksa varsayılan TRY (log için)
                logger.info(f"No header found, using default currency: TRY")
            elif color_category == 'ORANGE':  # Turuncu = İndirimli Fiyat
                column_mapping['discounted_price'] = col_idx
            # Diğer renkler veya renksiz veriler ignore edilir
        
        return column_mapping
    
    @staticmethod
    def _extract_products_from_sheet(sheet, header_row: int, column_mapping: Dict[str, int], company_name: str) -> List[Dict[str, Any]]:
        """Extract products from a sheet using column mapping"""
        import random
        products = []
        
        # Start row hesaplama: header varsa header_row + 1, yoksa 0
        start_row = 0 if header_row == -1 else header_row + 1
        
        # Start from the appropriate row
        for row_idx in range(start_row, sheet.max_row):
            try:
                # Extract data based on column mapping
                product_name = ""
                description = ""
                brand = ""  # Yeni: marka alanı
                detected_company = company_name  # Firma adı Excel yüklerken belirtilen firma
                list_price = 0
                discounted_price = None
                
                # Ürün adı (Kırmızı) - SADECE kırmızı hücre kabul edilir
                if column_mapping['product_name'] >= 0:
                    name_cell = sheet.cell(row=row_idx + 1, column=column_mapping['product_name'] + 1)
                    if (name_cell.value and 
                        ColorBasedExcelService.detect_color_category(name_cell.fill) == 'RED'):
                        product_name = str(name_cell.value).strip()

                # Açıklama (Mavi) - SADECE mavi hücre kabul edilir
                if column_mapping['description'] >= 0:
                    desc_cell = sheet.cell(row=row_idx + 1, column=column_mapping['description'] + 1)
                    if (desc_cell.value and 
                        ColorBasedExcelService.detect_color_category(desc_cell.fill) == 'BLUE'):
                        description = str(desc_cell.value).strip()

                # Marka (Sarı) - SADECE sarı hücre kabul edilir
                if column_mapping['brand'] >= 0:
                    brand_cell = sheet.cell(row=row_idx + 1, column=column_mapping['brand'] + 1)
                    if (brand_cell.value and 
                        ColorBasedExcelService.detect_color_category(brand_cell.fill) == 'YELLOW'):
                        brand_value = str(brand_cell.value).strip()
                        # Excel formülü değilse VE sayısal değer değilse kullan
                        if not brand_value.startswith('='):
                            try:
                                # Sayısal değer mi kontrol et
                                float(brand_value)
                                # Sayısal değerse boş bırak
                                logger.warning(f"Skipping numeric brand name: {brand_value}")
                            except ValueError:
                                # Sayısal değer değilse marka olarak kullan
                                brand = brand_value

                # Liste Fiyatı (Yeşil) - SADECE yeşil hücre kabul edilir
                if column_mapping['list_price'] >= 0:
                    price_cell = sheet.cell(row=row_idx + 1, column=column_mapping['list_price'] + 1)
                    if (price_cell.value and 
                        ColorBasedExcelService.detect_color_category(price_cell.fill) == 'GREEN'):
                        try:
                            list_price = float(str(price_cell.value).replace(',', '.'))
                        except:
                            list_price = 0

                # İndirimli Fiyat (Turuncu) - SADECE turuncu hücre kabul edilir
                if column_mapping['discounted_price'] >= 0:
                    disc_price_cell = sheet.cell(row=row_idx + 1, column=column_mapping['discounted_price'] + 1)
                    if (disc_price_cell.value and 
                        ColorBasedExcelService.detect_color_category(disc_price_cell.fill) == 'ORANGE'):
                        try:
                            discounted_price = float(str(disc_price_cell.value).replace(',', '.'))
                        except:
                            discounted_price = None

                # Özel mantık: Turuncu var ama yeşil yoksa
                if discounted_price and discounted_price > 0 and list_price == 0:
                    # İndirimli fiyat üzerine %20-30 arası rastgele zam ekle
                    markup_percentage = random.uniform(20, 30)
                    list_price = discounted_price * (1 + markup_percentage / 100)
                    logger.info(f"Generated list price for {product_name}: {discounted_price} + {markup_percentage:.1f}% = {list_price:.2f}")

                # Geçerli ürün kontrolü - en az liste fiyatı olmalı
                if (product_name and len(product_name) > 3 and 
                    list_price > 0 and 
                    not any(skip_word in product_name.lower() for skip_word in ['no', 'resim', 'ürün adı', 'toplam'])):
                    
                    products.append({
                        'name': product_name,
                        'description': description if description else None,
                        'brand': brand if brand else "",  # Yeni: marka alanı
                        'company_name': detected_company,
                        'list_price': list_price,
                        'discounted_price': discounted_price if discounted_price and discounted_price > 0 else None,
                        'currency': column_mapping.get('currency', 'TRY')  # Algılanan dövizi kullan
                    })
                    
            except Exception as e:
                logger.warning(f"Error processing row {row_idx + 1}: {e}")
                continue
        
        return products

# Excel parsing service
class ExcelService:
    @staticmethod
    def parse_excel_file(file_content: bytes) -> List[Dict[str, Any]]:
        """Parse Excel file and extract product data"""
        try:
            # Read Excel file
            df = pd.read_excel(io.BytesIO(file_content))
            
            logger.info(f"Excel file loaded: {len(df)} rows, {len(df.columns)} columns")
            
            # İlk veri satırını bul (header'ı tespit et)
            header_row = ExcelService._find_header_row(df)
            logger.info(f"Header row found at index: {header_row}")
            
            if header_row == -1:
                # Header bulunamazsa tüm kolonları kontrol et
                products = ExcelService._parse_without_header(df)
            else:
                # Header bulunduysa o satırdan itibaren parse et
                products = ExcelService._parse_with_header(df, header_row)
            
            logger.info(f"Total products extracted: {len(products)}")
            return products
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            raise HTTPException(status_code=400, detail=f"Excel dosyası işlenemedi: {str(e)}")
    
    @staticmethod
    def _find_header_row(df) -> int:
        """Excel dosyasında header satırını bul"""
        # Aranacak header kelimeleri
        header_keywords = [
            # Ürün adı varyantları
            'ürün', 'urun', 'product', 'malzeme', 'güneş', 'panel', 'solar', 'akü', 'batarya',
            'inverter', 'regülatör', 'kablo', 'malzemeler', 'items',
            # Fiyat varyantları  
            'fiyat', 'price', 'liste', 'list', 'tutar', 'amount', 'maliyet', 'cost',
            # İndirim varyantları
            'indirim', 'iskonto', 'discount', 'net', 'indirimli',
            # Para birimi varyantları
            'para', 'currency', 'birim', 'döviz', 'tl', 'usd', 'eur', '$', '€', '₺'
        ]
        
        for row_idx in range(min(20, len(df))):  # İlk 20 satırı kontrol et
            row_text = ""
            for col_idx in range(len(df.columns)):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value):
                    row_text += str(cell_value).lower() + " "
            
            # Bu satırda header keyword'leri var mı?
            keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
            if keyword_count >= 2:  # En az 2 keyword varsa header olabilir
                logger.info(f"Header row candidate at {row_idx}: '{row_text[:100]}...'")
                return row_idx
        
        return -1
    
    @staticmethod
    def _parse_with_header(df, header_row: int) -> List[Dict[str, Any]]:
        """Header'lı Excel dosyasını parse et"""
        # Header satırını kullanarak kolonları yeniden adlandır
        df_data = df.iloc[header_row:].copy()
        df_data.columns = df.iloc[header_row]
        df_data = df_data.iloc[1:]  # Header satırını atla
        
        return ExcelService._extract_products_from_dataframe(df_data)
    
    @staticmethod
    def _parse_without_header(df) -> List[Dict[str, Any]]:
        """Header'sız Excel dosyasını parse et"""
        logger.info("Parsing without header, analyzing all columns...")
        
        products = []
        
        # Her satırı kontrol et, veri olan satırları bul
        for row_idx in range(len(df)):
            row_data = {}
            product_name = ""
            list_price = 0
            discounted_price = None
            currency = "USD"  # Default currency
            
            # Bu satırdaki tüm hücreleri kontrol et
            for col_idx in range(len(df.columns)):
                cell_value = df.iloc[row_idx, col_idx]
                
                if pd.notna(cell_value):
                    str_value = str(cell_value).strip()
                    
                    # Fiyat hücresi mi kontrol et (sayısal değer)
                    try:
                        numeric_value = float(str_value.replace(',', '.'))
                        if 1 <= numeric_value <= 100000:  # Makul fiyat aralığı
                            if list_price == 0:
                                list_price = numeric_value
                            elif discounted_price is None and numeric_value < list_price:
                                discounted_price = numeric_value
                    except:
                        pass
                    
                    # Ürün adı hücresi mi kontrol et (text ve uzun)
                    if len(str_value) > 10 and any(char.isalpha() for char in str_value):
                        if not product_name or len(str_value) > len(product_name):
                            product_name = str_value
                    
                    # Para birimi kontrol et - gelişmiş algılama
                    detected_currency = ExcelService.detect_currency_from_text(str_value, None)
                    if detected_currency:
                        currency = detected_currency
            
            # Geçerli ürün bilgisi var mı kontrol et
            if product_name and list_price > 0:
                products.append({
                    'name': product_name,
                    'list_price': list_price,
                    'currency': currency,
                    'discounted_price': discounted_price
                })
                logger.info(f"Extracted product: {product_name[:50]}... - ${list_price}")
        
        return products
    
    @staticmethod
    def _extract_products_from_dataframe(df) -> List[Dict[str, Any]]:
        """DataFrame'den ürün verilerini çıkar"""
        products = []
        
        # ELEKTROZİRVE formatını kontrol et (basit 4 kolon)
        if len(df.columns) == 4:
            logger.info("Detected ELEKTROZİRVE format (4 columns)")
            return ExcelService._parse_elektrozirve_format(df)
        
        # HAVENSİS formatını kontrol et (karmaşık multi-column)
        elif len(df.columns) > 10:
            logger.info("Detected HAVENSİS format (multi-column)")
            return ExcelService._parse_havensis_format(df)
        
        # Genel format parsing
        else:
            logger.info("Using general format parsing")
            return ExcelService._parse_general_format(df)
    
    @staticmethod
    def _parse_elektrozirve_format(df) -> List[Dict[str, Any]]:
        """ELEKTROZİRVE formatında Excel parse et"""
        products = []
        
        # İlk satır header (Güneş Panelleri, LİSTE FİYATI, İskonto, Net Fiyat)
        df.columns = ['product_name', 'list_price', 'discount_rate', 'net_price']
        
        for index, row in df.iterrows():
            try:
                if index == 0:  # Header satırını atla
                    continue
                
                product_name = str(row['product_name']).strip() if pd.notna(row['product_name']) else ""
                list_price = float(row['list_price']) if pd.notna(row['list_price']) else 0
                net_price = float(row['net_price']) if pd.notna(row['net_price']) else 0
                
                # Kategori başlıkları atla (örn: "Esnek Güneş Panelleri")
                if ('panelleri' in product_name.lower() or 'aküler' in product_name.lower() or 
                    'regülatörler' in product_name.lower()) and list_price == 0:
                    continue
                
                # Geçerli ürün kontrolü
                if (product_name and len(product_name) > 5 and 
                    list_price > 0 and not product_name.lower().startswith('liste')):
                    
                    products.append({
                        'name': product_name,
                        'list_price': list_price,
                        'currency': 'TRY',  # ELEKTROZİRVE TL fiyatları
                        'discounted_price': net_price if net_price != list_price else None
                    })
                    logger.info(f"Added ELEKTROZİRVE product: {product_name[:50]}... - {list_price} TL")
                    
            except Exception as e:
                logger.warning(f"Error processing ELEKTROZİRVE row {index}: {e}")
                continue
        
        return products
    
    @staticmethod
    def _parse_havensis_format(df) -> List[Dict[str, Any]]:
        """HAVENSİS formatında Excel parse et"""
        products = []
        
        # HAVENSİS formatı: Col3=Ürün, Col6=Fiyat$, Col7=İskonto, Col8=İskontolu Fiyat$
        product_col = 3
        price_col = 6
        discount_rate_col = 7
        discounted_price_col = 8
        
        for index, row in df.iterrows():
            try:
                if index < 12:  # İlk 12 satır header/boş satırlar
                    continue
                
                # Ürün adı (Col3)
                product_name = ""
                if len(row) > product_col and pd.notna(row.iloc[product_col]):
                    product_name = str(row.iloc[product_col]).strip()
                
                # Liste fiyatı (Col6)
                list_price = 0
                if len(row) > price_col and pd.notna(row.iloc[price_col]):
                    try:
                        list_price = float(row.iloc[price_col])
                    except:
                        list_price = 0
                
                # İndirimli fiyat (Col8)
                discounted_price = None
                if len(row) > discounted_price_col and pd.notna(row.iloc[discounted_price_col]):
                    try:
                        discounted_price = float(row.iloc[discounted_price_col])
                    except:
                        discounted_price = None
                
                # Geçerli ürün kontrolü
                if (product_name and len(product_name) > 10 and 
                    list_price > 0 and 'panel' in product_name.lower()):
                    
                    products.append({
                        'name': product_name,
                        'list_price': list_price,
                        'currency': 'USD',  # HAVENSİS USD fiyatları
                        'discounted_price': discounted_price
                    })
                    logger.info(f"Added HAVENSİS product: {product_name[:50]}... - ${list_price}")
                    
            except Exception as e:
                logger.warning(f"Error processing HAVENSİS row {index}: {e}")
                continue
        
        return products
    
    @staticmethod
    def detect_currency_from_text(text: str, fallback_currency: str = 'USD') -> str:
        """Detect currency from any text (header, cell value, etc.)"""
        if not text:
            return fallback_currency
            
        text = str(text).upper().strip()
        
        # Dolar kontrolü - gelişmiş
        dollar_keywords = ['$', 'DOLAR', 'DOLLAR', 'USD', 'DOLAR İSARETİ', 'DOLAR IŞARETI', 'AMERİKAN DOLARI', 'AMERIKAN DOLARI']
        if any(keyword in text for keyword in dollar_keywords):
            return 'USD'
        
        # Euro kontrolü - gelişmiş
        euro_keywords = ['€', 'EURO', 'EUR', 'AVRO', 'AVRUPA']
        if any(keyword in text for keyword in euro_keywords):
            return 'EUR'
        
        # TL kontrolü - gelişmiş
        tl_keywords = ['₺', 'TL', 'TRY', 'TÜRK', 'LIRA', 'TÜRK LİRASI', 'TURK LIRASI', 'TURKİYE', 'TURKIYE']
        if any(keyword in text for keyword in tl_keywords):
            return 'TRY'
            
        return fallback_currency

    @staticmethod
    def _parse_general_format(df) -> List[Dict[str, Any]]:
        """Genel format parsing"""
        products = []
        
        # Gelişmiş kolon mapping
        column_mapping = {
            # Ürün adı varyantları
            'ürün adı': 'product_name', 'urun adi': 'product_name', 'product name': 'product_name',
            'ürün': 'product_name', 'urun': 'product_name', 'product': 'product_name',
            'malzeme': 'product_name', 'malzemeler': 'product_name', 'item': 'product_name',
            'güneş panelleri': 'product_name', 'gunes panelleri': 'product_name',
            'solar panel': 'product_name', 'panel': 'product_name',
            'aküler': 'product_name', 'akü': 'product_name', 'batarya': 'product_name',
            'ad': 'product_name', 'name': 'product_name',
            
            # Marka varyantları (yeni)
            'marka': 'brand', 'brand': 'brand', 'markalar': 'brand', 'brands': 'brand',
            'üretici': 'brand', 'uretici': 'brand', 'manufacturer': 'brand',
            'yapimci': 'brand', 'yapımcı': 'brand', 'maker': 'brand',
            
            # Liste fiyatı varyantları
            'liste fiyatı': 'list_price', 'liste fiyati': 'list_price', 'list price': 'list_price',
            'fiyat$': 'list_price', 'fiyat $': 'list_price', 'fiyat': 'list_price',
            'liste': 'list_price', 'list': 'list_price', 'price': 'list_price',
            'tutar': 'list_price', 'amount': 'list_price', 'maliyet': 'list_price',
            
            # İndirimli fiyat varyantları
            'indirimli fiyat $': 'discounted_price', 'indirimli fiyat$': 'discounted_price',
            'iskontolu fiyat $': 'discounted_price', 'iskontolu fiyat$': 'discounted_price',
            'indirimli fiyat': 'discounted_price', 'indirimli fiyati': 'discounted_price',
            'iskontolu fiyat': 'discounted_price', 'iskontolu fiyati': 'discounted_price',
            'net fiyat': 'discounted_price', 'net price': 'discounted_price',
            'discounted price': 'discounted_price', 'discount price': 'discounted_price',
            'net': 'discounted_price', 'indirim': 'discounted_price', 'iskonto': 'discounted_price',
            
            # Para birimi varyantları
            'para birimi': 'currency', 'currency': 'currency', 'birim': 'currency',
            'döviz': 'currency', 'doviz': 'currency'
        }
        
        # Kolonları normalize et
        df.columns = df.columns.astype(str).str.lower().str.strip()
        logger.info(f"Normalized columns: {list(df.columns)}")
        
        # Kolon mapping uygula
        for col in df.columns:
            for mapping_key, mapping_value in column_mapping.items():
                if mapping_key in col:
                    df = df.rename(columns={col: mapping_value})
                    logger.info(f"Mapped column '{col}' to '{mapping_value}'")
                    break
        
        logger.info(f"Final mapped columns: {list(df.columns)}")
        
        # Eğer standart kolonlar yoksa, konum bazlı mapping dene
        if 'product_name' not in df.columns:
            # İlk metin kolonu ürün adı olabilir
            for col in df.columns:
                if df[col].dtype == 'object':
                    df = df.rename(columns={col: 'product_name'})
                    logger.info(f"Using first text column '{col}' as product_name")
                    break
        
        if 'list_price' not in df.columns:
            # İlk sayısal kolon liste fiyatı olabilir  
            for col in df.columns:
                if col != 'product_name' and pd.api.types.is_numeric_dtype(df[col]):
                    df = df.rename(columns={col: 'list_price'})
                    logger.info(f"Using first numeric column '{col}' as list_price")
                    break
        
        # Ürünleri çıkar
        for index, row in df.iterrows():
            try:
                # Ürün adı
                product_name = ""
                if 'product_name' in row:
                    product_name = str(row['product_name']).strip() if pd.notna(row['product_name']) else ""
                
                # Marka (yeni)
                brand = ""
                if 'brand' in row and pd.notna(row['brand']):
                    brand = str(row['brand']).strip()
                
                # Liste fiyatı
                list_price = 0
                if 'list_price' in row:
                    try:
                        list_price = float(row['list_price']) if pd.notna(row['list_price']) else 0
                    except:
                        list_price = 0
                
                # İndirimli fiyat
                discounted_price = None
                if 'discounted_price' in row and pd.notna(row['discounted_price']):
                    try:
                        discounted_price = float(row['discounted_price'])
                    except:
                        discounted_price = None
                
                # Para birimi algılama - gelişmiş
                currency = "USD"  # varsayılan
                
                # Önce currency sütunu varsa onu kullan
                if 'currency' in row and pd.notna(row['currency']):
                    detected = ExcelService.detect_currency_from_text(str(row['currency']))
                    if detected:
                        currency = detected
                
                # Tüm sütunlarda para birimi işaretçilerini ara
                for col_name, col_value in row.items():
                    if pd.notna(col_value):
                        cell_text = str(col_value)
                        detected = ExcelService.detect_currency_from_text(cell_text, None)
                        if detected:
                            currency = detected
                            break
                
                # Sütun başlıklarında da para birimi ara
                for col_name in df.columns:
                    detected = ExcelService.detect_currency_from_text(str(col_name), None)
                    if detected:
                        currency = detected
                        break
                
                # Geçerli ürün kontrolü
                if product_name and len(product_name) > 3 and list_price > 0:
                    product = {
                        'name': product_name,
                        'brand': brand,  # Yeni: marka alanı
                        'list_price': list_price,
                        'currency': currency,
                        'discounted_price': discounted_price
                    }
                    products.append(product)
                    logger.info(f"Added product: {product_name[:50]}... - {list_price} {currency}")
                else:
                    logger.info(f"Skipped row {index}: name='{product_name}', price={list_price}")
                    
            except Exception as e:
                logger.warning(f"Error processing row {index}: {e}")
                continue
        
        return products

excel_service = ExcelService()

# API Routes

@api_router.get("/")
async def root():
    return {"message": "Karavan Elektrik Ekipmanları Fiyat Karşılaştırma API"}

@api_router.get("/exchange-rates")
async def get_exchange_rates():
    """Get current exchange rates"""
    try:
        rates = await currency_service.get_exchange_rates()
        return {
            "success": True,
            "rates": {k: float(v) for k, v in rates.items()},
            "updated_at": currency_service.last_update.isoformat() if currency_service.last_update else None
        }
    except Exception as e:
        logger.error(f"Error getting exchange rates: {e}")
        raise HTTPException(status_code=500, detail="Döviz kurları alınamadı")

@api_router.post("/exchange-rates/update")
async def update_exchange_rates():
    """Force update exchange rates from API"""
    try:
        # Clear cache to force fresh API call
        currency_service.rates_cache = {}
        currency_service.last_update = None
        
        rates = await currency_service.get_exchange_rates()
        return {
            "success": True,
            "message": "Döviz kurları başarıyla güncellendi",
            "rates": {k: float(v) for k, v in rates.items()},
            "updated_at": currency_service.last_update.isoformat() if currency_service.last_update else None
        }
    except Exception as e:
        logger.error(f"Error updating exchange rates: {e}")
        raise HTTPException(status_code=500, detail="Döviz kurları güncellenemedi")

@api_router.post("/companies", response_model=Company)
async def create_company(company: CompanyCreate):
    """Create a new company"""
    try:
        company_dict = {
            "id": str(uuid.uuid4()),
            "name": company.name,
            "created_at": datetime.now(timezone.utc)
        }
        
        result = await db.companies.insert_one(company_dict)
        return Company(**company_dict)
        
    except Exception as e:
        logger.error(f"Error creating company: {e}")
        raise HTTPException(status_code=500, detail="Firma oluşturulamadı")

@api_router.get("/companies", response_model=List[Company])
async def get_companies():
    """Get all companies"""
    try:
        companies = await db.companies.find().to_list(None)
        return [Company(**company) for company in companies]
    except Exception as e:
        logger.error(f"Error getting companies: {e}")
        raise HTTPException(status_code=500, detail="Firmalar getirilemedi")

@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str):
    """Delete a company"""
    try:
        result = await db.companies.delete_one({"id": company_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Firma bulunamadı")
        
        # Also delete all products of this company
        await db.products.delete_many({"company_id": company_id})
        
        return {"success": True, "message": "Firma silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting company: {e}")
        raise HTTPException(status_code=500, detail="Firma silinemedi")

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    brand: Optional[str] = None  # Marka alanı
    company_id: Optional[str] = None  # Firma alanı
    description: Optional[str] = None
    image_url: Optional[str] = None
    list_price: Optional[Decimal] = None
    discounted_price: Optional[Decimal] = None
    currency: Optional[str] = None
    category_id: Optional[str] = None

@api_router.patch("/products/{product_id}")
async def update_product(product_id: str, update_data: ProductUpdate):
    """Update a product"""
    try:
        # Get existing product
        existing_product = await db.products.find_one({"id": product_id})
        if not existing_product:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Prepare update data
        update_dict = {}
        if update_data.name is not None:
            update_dict["name"] = update_data.name
        if update_data.brand is not None:
            update_dict["brand"] = update_data.brand
        if update_data.company_id is not None:
            update_dict["company_id"] = update_data.company_id
        if update_data.description is not None:
            update_dict["description"] = update_data.description
        if update_data.image_url is not None:
            update_dict["image_url"] = update_data.image_url
        if update_data.list_price is not None:
            update_dict["list_price"] = float(update_data.list_price)
        if update_data.discounted_price is not None:
            update_dict["discounted_price"] = float(update_data.discounted_price)
        if update_data.currency is not None:
            update_dict["currency"] = update_data.currency.upper()
        if update_data.category_id is not None:
            update_dict["category_id"] = update_data.category_id
        
        # If currency or prices changed, recalculate TRY prices
        if update_data.currency is not None or update_data.list_price is not None or update_data.discounted_price is not None:
            currency = update_data.currency.upper() if update_data.currency else existing_product["currency"]
            list_price = float(update_data.list_price) if update_data.list_price is not None else existing_product["list_price"]
            discounted_price = float(update_data.discounted_price) if update_data.discounted_price is not None else existing_product.get("discounted_price")
            
            # Convert to TRY
            try:
                list_price_try = await currency_service.convert_to_try(Decimal(str(list_price)), currency)
                update_dict["list_price_try"] = float(list_price_try)
            except Exception as e:
                logger.warning(f"Failed to convert list price to TRY: {e}")
                update_dict["list_price_try"] = float(list_price)
            
            if discounted_price is not None:
                try:
                    discounted_price_try = await currency_service.convert_to_try(Decimal(str(discounted_price)), currency)
                    update_dict["discounted_price_try"] = float(discounted_price_try)
                except Exception as e:
                    logger.warning(f"Failed to convert discounted price to TRY: {e}")
                    update_dict["discounted_price_try"] = float(discounted_price)
            else:
                update_dict["discounted_price_try"] = None
        
        # Update product
        if update_dict:
            result = await db.products.update_one(
                {"id": product_id},
                {"$set": update_dict}
            )
            
            if result.modified_count == 0:
                raise HTTPException(status_code=404, detail="Ürün güncellenemedi")
        
        # Get updated product
        updated_product = await db.products.find_one({"id": product_id})
        return {
            "success": True,
            "message": "Ürün başarıyla güncellendi",
            "product": Product(**updated_product)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating product: {e}")
        raise HTTPException(status_code=500, detail="Ürün güncellenemedi")

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    """Delete a product"""
    try:
        result = await db.products.delete_one({"id": product_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        return {"success": True, "message": "Ürün silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting product: {e}")
        raise HTTPException(status_code=500, detail="Ürün silinemedi")

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, product_update: Dict[str, Any]):
    """Update a product (especially for category assignment)"""
    try:
        # Mevcut ürünü bul
        existing_product = await db.products.find_one({"id": product_id})
        if not existing_product:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Güncellenecek alanları hazırla
        update_data = {}
        
        # Kategori güncellenmesi
        if "category_id" in product_update:
            category_id = product_update["category_id"]
            
            # Kategori var mı kontrol et (eğer none değilse)
            if category_id and category_id != "none":
                category = await db.categories.find_one({"id": category_id})
                if not category:
                    raise HTTPException(status_code=404, detail="Kategori bulunamadı")
            
            update_data["category_id"] = category_id
        
        # Diğer güncellenebilir alanlar
        allowed_fields = ["name", "brand", "description", "list_price", "discounted_price", "currency", "company_id"]
        for field in allowed_fields:
            if field in product_update:
                update_data[field] = product_update[field]
        
        # Güncelleme zamanını ekle
        update_data["updated_at"] = datetime.utcnow().isoformat() + "Z"
        
        # Ürünü güncelle
        result = await db.products.update_one(
            {"id": product_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Güncellenmiş ürünü döndür
        updated_product = await db.products.find_one({"id": product_id})
        if updated_product:
            updated_product.pop('_id', None)
        
        return updated_product
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating product: {e}")
        raise HTTPException(status_code=500, detail=f"Ürün güncellenemedi: {str(e)}")

# ===== QUOTE ENDPOINTS =====

@api_router.post("/quotes", response_model=QuoteResponse)
async def create_quote(quote: QuoteCreate):
    """Create a new quote"""
    try:
        # Validate products exist
        product_ids = [p["id"] for p in quote.products]
        products_cursor = db.products.find({"id": {"$in": product_ids}})
        products = await products_cursor.to_list(length=None)
        
        if len(products) != len(quote.products):
            raise HTTPException(status_code=400, detail="Some products not found")
        
        # Create product-quantity mapping
        product_quantities = {p["id"]: p.get("quantity", 1) for p in quote.products}
        
        # Calculate totals
        total_list_price = 0
        total_discounted_price = 0
        processed_products = []
        
        for product in products:
            # Get company info
            company = await db.companies.find_one({"id": product["company_id"]})
            
            # Get quantity for this product
            quantity = product_quantities.get(product["id"], 1)
            
            list_price_try = float(product.get("list_price_try", 0))
            discounted_price_try = float(product.get("discounted_price_try", 0)) if product.get("discounted_price_try") else list_price_try
            
            # Calculate totals with quantity
            total_list_price += list_price_try * quantity
            total_discounted_price += discounted_price_try * quantity
            
            processed_products.append({
                "id": product["id"],
                "name": product["name"],
                "description": product.get("description"),
                "company_name": company["name"] if company else "Unknown",
                "list_price": product["list_price"],
                "list_price_try": list_price_try,
                "discounted_price": product.get("discounted_price"),
                "discounted_price_try": discounted_price_try,
                "currency": product["currency"],
                "quantity": quantity
            })
        
        # Apply quote discount
        quote_discount_amount = total_discounted_price * (quote.discount_percentage / 100)
        
        # Add labor cost to the calculation
        labor_cost = float(quote.labor_cost)
        total_with_labor = total_discounted_price - quote_discount_amount + labor_cost
        total_net_price = total_with_labor
        
        # Create quote document
        quote_doc = {
            "id": str(uuid.uuid4()),
            "name": quote.name,
            "customer_name": quote.customer_name,
            "customer_email": quote.customer_email,
            "discount_percentage": quote.discount_percentage,
            "labor_cost": labor_cost,  # İşçilik maliyeti eklendi
            "total_list_price": total_list_price,
            "total_discounted_price": total_discounted_price,
            "total_net_price": total_net_price,
            "products": processed_products,
            "notes": quote.notes,
            "created_at": datetime.utcnow().isoformat() + "Z",
            "status": "active"
        }
        
        result = await db.quotes.insert_one(quote_doc)
        
        logger.info(f"Quote created: {quote.name} with {len(products)} products")
        return quote_doc
        
    except Exception as e:
        logger.error(f"Error creating quote: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating quote: {str(e)}")

@api_router.get("/quotes", response_model=List[QuoteResponse])
async def get_quotes():
    """Get all quotes"""
    try:
        quotes_cursor = db.quotes.find({"status": "active"}).sort("created_at", -1)
        quotes = await quotes_cursor.to_list(length=None)
        
        return quotes
        
    except Exception as e:
        logger.error(f"Error fetching quotes: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching quotes: {str(e)}")

@api_router.get("/quotes/{quote_id}", response_model=QuoteResponse)
async def get_quote(quote_id: str):
    """Get specific quote by ID"""
    try:
        quote = await db.quotes.find_one({"id": quote_id})
        
        if not quote:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        return quote
        
    except Exception as e:
        logger.error(f"Error fetching quote: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching quote: {str(e)}")

@api_router.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: str):
    """Delete a quote"""
    try:
        result = await db.quotes.update_one(
            {"id": quote_id},
            {"$set": {"status": "deleted"}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        return {"success": True, "message": "Quote deleted successfully"}
        
    except Exception as e:
        logger.error(f"Error deleting quote: {e}")
        raise HTTPException(status_code=500, detail=f"Error deleting quote: {str(e)}")

@api_router.put("/quotes/{quote_id}")
async def update_quote(quote_id: str, quote_update: Dict[str, Any]):
    """Update an existing quote (especially for adding labor cost)"""
    try:
        # Mevcut teklifi bul
        existing_quote = await db.quotes.find_one({"id": quote_id, "status": "active"})
        if not existing_quote:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        # Güncellenecek alanları hazırla
        update_data = {}
        
        # İşçilik maliyeti güncellenirse toplam hesapla
        if "labor_cost" in quote_update:
            labor_cost = float(quote_update["labor_cost"])
            
            # Mevcut hesapları al
            total_discounted_price = existing_quote.get("total_discounted_price", 0)
            discount_percentage = existing_quote.get("discount_percentage", 0)
            
            # Net toplam hesapla
            discount_amount = total_discounted_price * (discount_percentage / 100)
            total_net_price = total_discounted_price - discount_amount + labor_cost
            
            update_data["labor_cost"] = labor_cost
            update_data["total_net_price"] = total_net_price
        
        # Diğer alanları da güncelle
        if "discount_percentage" in quote_update:
            update_data["discount_percentage"] = float(quote_update["discount_percentage"])
        
        # Güncelleme zamanını ekle
        update_data["updated_at"] = datetime.utcnow().isoformat() + "Z"
        
        # Teklifi güncelle
        result = await db.quotes.update_one(
            {"id": quote_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        # Güncellenmiş teklifi döndür
        updated_quote = await db.quotes.find_one({"id": quote_id})
        if updated_quote:
            updated_quote.pop('_id', None)
        
        return updated_quote
        
    except Exception as e:
        logger.error(f"Error updating quote: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating quote: {str(e)}")

class PDFQuoteGenerator:
    def __init__(self):
        self.setup_fonts()
        self.styles = getSampleStyleSheet()
        self.setup_styles()
    
    def setup_fonts(self):
        """Türkçe karakter desteği için font kurulumu"""
        self.montserrat_available = False
        self.montserrat_bold_available = False
        
        try:
            font_dir = Path(__file__).parent / 'fonts'
            
            # Montserrat Regular font
            montserrat_regular_path = font_dir / 'Montserrat-Regular.ttf'
            if montserrat_regular_path.exists():
                # Unicode subset ile kaydet
                pdfmetrics.registerFont(TTFont('Montserrat', str(montserrat_regular_path), subfontIndex=0))
                self.montserrat_available = True
                logger.info("Montserrat Regular font loaded successfully")
                
                # Test için Türkçe karakter desteğini kontrol et
                from reportlab.pdfbase.pdfmetrics import getFont
                try:
                    font = getFont('Montserrat')
                    # Türkçe karakterleri test et
                    test_chars = 'ğüşıöç ĞÜŞIÖÇ'
                    for char in test_chars:
                        if hasattr(font, 'face') and hasattr(font.face, 'getCharWidth'):
                            width = font.face.getCharWidth(ord(char))
                            if width == 0:
                                logger.warning(f"Character '{char}' not supported in Montserrat")
                except Exception as e:
                    logger.warning(f"Font test failed: {e}")
            else:
                logger.warning("Montserrat Regular font not found")
            
            # Montserrat Bold font
            montserrat_bold_path = font_dir / 'Montserrat-Bold.ttf'
            if montserrat_bold_path.exists():
                pdfmetrics.registerFont(TTFont('Montserrat-Bold', str(montserrat_bold_path), subfontIndex=0))
                self.montserrat_bold_available = True
                logger.info("Montserrat Bold font loaded successfully")
            else:
                logger.warning("Montserrat Bold font not found")
                
        except Exception as e:
            logger.error(f"Font loading error: {e}")
            self.montserrat_available = False
            self.montserrat_bold_available = False
    
    def get_font_name(self, is_bold=False):
        """Montserrat font adını döndür - Türkçe karakter desteği ile"""
        if is_bold and self.montserrat_bold_available:
            return 'Montserrat-Bold'
        elif not is_bold and self.montserrat_available:
            return 'Montserrat'
        elif is_bold:
            return 'Helvetica-Bold'
        else:
            return 'Helvetica'
    
    def setup_styles(self):
        """PDF için özel stiller tanımla - Yeni renk şeması ile"""
        
        # Yeni renk paleti
        primary_color = colors.HexColor('#25c7eb')      # Tablo renkleri için
        secondary_color = colors.HexColor('#1ba3cc')    # Eski renkler için
        new_primary_color = colors.HexColor('#2F4B68')  # Ana başlıklar için YENİ
        table_header_color = colors.HexColor('#A6C9EC') # Tablo başlık arka planı YENİ
        accent_color = colors.HexColor('#85e8ff')       # Açık turkuaz
        text_color = colors.HexColor('#2d3748')         # Koyu gri
        
        # Başlık stili - Yeni renk (#2F4B68)
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontName=self.get_font_name(is_bold=True),
            fontSize=22,
            spaceAfter=25,
            spaceBefore=10,
            alignment=TA_CENTER,
            textColor=new_primary_color,  # YENİ RENK
            leading=26
        )
        
        # Alt başlık stili - Yeni renk (#2F4B68)
        self.subtitle_style = ParagraphStyle(
            'SubTitle',
            parent=self.styles['Heading2'],
            fontName=self.get_font_name(is_bold=True),
            fontSize=14,
            spaceAfter=15,
            spaceBefore=10,
            alignment=TA_LEFT,
            textColor=new_primary_color,  # YENİ RENK
            leading=18
        )
        
        # Firma bilgi stili
        self.company_style = ParagraphStyle(
            'CompanyInfo',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(),
            fontSize=11,
            spaceAfter=6,
            alignment=TA_LEFT,
            textColor=text_color,
            leading=14
        )
        
        # Normal metin stili
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(),
            fontSize=10,
            spaceAfter=8,
            textColor=text_color,
            leading=13
        )
        
        # Veri stili (tablolar için)
        self.data_style = ParagraphStyle(
            'DataText',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(),
            fontSize=9,
            spaceAfter=4,
            textColor=text_color,
            leading=12
        )
        
        # Footer stili
        self.footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(),
            fontSize=9,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#718096'),
            spaceAfter=6,
            leading=11
        )
        
        # Fiyat vurgu stili - Küçültülmüş ve yeni renk (#2F4B68)
        self.price_style = ParagraphStyle(
            'PriceHighlight',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(is_bold=True),
            fontSize=14,  # 16'dan 14'e küçültüldü
            alignment=TA_RIGHT,
            textColor=new_primary_color,  # YENİ RENK
            spaceAfter=10,
            leading=18
        )

    def create_quote_pdf(self, quote_data: Dict) -> BytesIO:
        """Gelişmiş teklif PDF'i oluştur - Türkçe karakter desteği ile"""
        buffer = BytesIO()
        
        # Yüksek kaliteli PDF ayarları
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2.5*cm,
            leftMargin=2.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
            title=f"Teklif - {quote_data.get('name', 'Adsız')}",
            author="Karavan Elektrik Ekipmanları"
        )
        
        story = []
        
        # Üst header bölümü
        story.append(self._create_modern_header())
        story.append(Spacer(1, 25))
        
        # Teklif başlığı
        quote_name = quote_data.get('name', 'Fiyat Teklifi')
        story.append(Paragraph(f"<b>{quote_name}</b>", self.title_style))
        story.append(Spacer(1, 15))
        
        # Teklif bilgileri satırı (Tarih, Geçerlilik vs.)
        story.append(self._create_quote_info_section(quote_data))
        story.append(Spacer(1, 25))
        
        # Ürün tablosu başlığı
        story.append(Paragraph("<b>Teklif Detayları</b>", self.subtitle_style))
        story.append(Spacer(1, 10))
        
        # Ürün tablosu
        story.append(self._create_modern_products_table(quote_data['products']))
        story.append(Spacer(1, 25))
        
        # Toplam hesaplama bölümü
        story.extend(self._create_modern_totals_section(quote_data))
        story.append(Spacer(1, 30))
        
        # Footer notları
        story.extend(self._create_modern_footer())
        
        # PDF oluştur
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _create_modern_header(self):
        """Logo ve Çorlu Karavan bilgileri başlığı - Yeni logo ile"""
        from reportlab.platypus import Table as PDFTable
        
        # Yeni logo yolu
        logo_path = Path(__file__).parent / 'images' / 'corlu_karavan_logo_new.png'
        
        # Header tablosu oluştur (Logo + Firma bilgileri)
        if logo_path.exists():
            try:
                logo_img = Image(str(logo_path), width=80, height=80)
                
                # Firma bilgileri - ÇORLU KARAVAN yeni renk ile (#2F4B68)
                company_info = [
                    "<font size='16' color='#2F4B68'><b>ÇORLU KARAVAN</b></font>",
                    " ",
                    "<font size='10'>Adres: Hatip, Sarı Salkım 3.Sokak Mobilyacılar Sitesi No: B1, 59000 Çorlu/Tekirdağ</font>",
                    "<font size='10'>Telefon: 0505 813 77 65</font>",
                    "<font size='10'>E-posta: info@corlukaravan.com</font>",
                    "<font size='10'>Teknik Destek: mehmetnecdet@corlukaravan.com</font>"
                ]
                
                company_text = "<br/>".join(company_info)
                company_paragraph = Paragraph(company_text, self.company_style)
                
                # Logo ve metin tablosu
                header_data = [[logo_img, company_paragraph]]
                header_table = PDFTable(header_data, colWidths=[100, 400])
                header_table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo ortala
                    ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Metin sola hizala
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Dikey ortala
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ]))
                
                return header_table
                
            except Exception as e:
                logger.error(f"Logo loading error: {e}")
        
        # Logo yoksa sadece metin - ÇORLU KARAVAN yeni renk ile
        company_info = [
            "<font size='16' color='#2F4B68'><b>ÇORLU KARAVAN</b></font>",
            " ",
            "<font size='10'>Adres: Hatip, Sarı Salkım 3.Sokak Mobilyacılar Sitesi No: B1, 59000 Çorlu/Tekirdağ</font>",
            "<font size='10'>Telefon: 0505 813 77 65</font>",
            "<font size='10'>E-posta: info@corlukaravan.com</font>",
            "<font size='10'>Teknik Destek: mehmetnecdet@corlukaravan.com</font>"
        ]
        
        header_text = "<br/>".join(company_info)
        return Paragraph(header_text, self.company_style)
    
    def _create_quote_info_section(self, quote_data: Dict):
        """Teklif bilgileri bölümü (başlangıç tarihi ve 7 gün sonrası bitiş tarihi)"""
        try:
            created_date = datetime.fromisoformat(quote_data['created_at'].replace('Z', '+00:00'))
            start_date_str = created_date.strftime('%d.%m.%Y')
            
            # Bitiş tarihi her zaman 7 gün sonra
            end_date = created_date + timedelta(days=7)
            end_date_str = end_date.strftime('%d.%m.%Y')
        except:
            today = datetime.now()
            start_date_str = today.strftime('%d.%m.%Y')
            end_date_str = (today + timedelta(days=7)).strftime('%d.%m.%Y')
        
        # Bilgi tablosu oluştur
        info_data = [
            [
                Paragraph("<b>Başlangıç Tarihi:</b>", self.normal_style),
                Paragraph(start_date_str, self.normal_style),
                Paragraph("<b>Bitiş Tarihi:</b>", self.normal_style),
                Paragraph(end_date_str, self.normal_style)
            ]
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 3*cm, 4*cm, 3*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f7fafc')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), self.get_font_name()),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        return info_table
    
    def _create_date_header(self):
        """Sağ üst köşede tarih"""
        from datetime import datetime
        from reportlab.platypus import Table as PDFTable
        
        # Tarih
        now = datetime.now()
        date_str = now.strftime('%d.%m.%Y')
        
        # Sağa hizalı tarih
        date_style = ParagraphStyle(
            'DateHeader',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(),
            fontSize=11,
            alignment=TA_RIGHT,
            textColor=colors.HexColor('#4A5568')
        )
        
        date_paragraph = Paragraph(f"Tarih: {date_str}", date_style)
        return date_paragraph
    
    def _create_modern_products_table(self, products: List[Dict]):
        """Modern tasarımda ürün tablosu oluştur - Yeni renk şeması ile"""
        primary_color = colors.HexColor('#25c7eb')
        secondary_color = colors.HexColor('#1ba3cc')
        table_header_color = colors.HexColor('#A6C9EC')  # YENİ TABLO BAŞLIK RENGİ
        accent_color = colors.HexColor('#f0f9ff')
        
        # Tablo başlıkları (Marka sütunu kaldırıldı)
        headers = [
            Paragraph("<b>Ürün Adı</b>", self.data_style),
            Paragraph("<b>Miktar</b>", self.data_style),
            Paragraph("<b>Birim Fiyat</b>", self.data_style),
            Paragraph("<b>Toplam Fiyat</b>", self.data_style)
        ]
        
        data = [headers]
        
        # Ürün satırları (Marka sütunu kaldırıldı)
        for product in products:
            quantity = product.get('quantity', 1)
            unit_price = product.get('discounted_price_try', product.get('list_price_try', 0))
            total_price = unit_price * quantity
            
            row = [
                Paragraph(product.get('name', ''), self.data_style),
                Paragraph(str(quantity), self.data_style),
                Paragraph(f"₺ {self._format_price_modern(unit_price)}", self.data_style),
                Paragraph(f"<b>₺ {self._format_price_modern(total_price)}</b>", self.data_style)
            ]
            data.append(row)
        
        # Tablo oluştur (4 sütun)
        table = Table(data, colWidths=[8*cm, 2*cm, 3*cm, 3*cm])
        table.setStyle(TableStyle([
            # Başlık stili - YENİ RENK (#A6C9EC)
            ('BACKGROUND', (0, 0), (-1, 0), table_header_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.get_font_name(is_bold=True)),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Veri satırları - Alternatif renklendirme (4 sütun)
            ('BACKGROUND', (0, 1), (-1, -1), accent_color),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, accent_color]),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Miktar ortala
            ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Fiyatları sağa hizala
            ('FONTNAME', (0, 1), (-1, -1), self.get_font_name()),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Çerçeve
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Son sütun vurgusu (toplam fiyatlar için)
            ('FONTNAME', (3, 1), (3, -1), self.get_font_name(is_bold=True)),
            ('TEXTCOLOR', (3, 1), (3, -1), secondary_color),
        ]))
        
        return table
    
    def _create_modern_totals_section(self, quote_data: Dict):
        """Modern toplam hesaplama bölümü - İşçilik maliyeti ile"""
        primary_color = colors.HexColor('#25c7eb')
        secondary_color = colors.HexColor('#1ba3cc')
        new_primary_color = colors.HexColor('#2F4B68')
        
        totals_content = []
        
        # Ara toplam
        subtotal_text = f"Ara Toplam: <b>₺ {self._format_price_modern(quote_data.get('total_discounted_price', 0))}</b>"
        totals_content.append(Paragraph(subtotal_text, self.normal_style))
        
        # İndirim (eğer varsa)  
        discount_percentage = quote_data.get('discount_percentage', 0)
        if discount_percentage > 0:
            discount_amount = quote_data.get('total_discounted_price', 0) - quote_data.get('total_net_price', 0) + quote_data.get('labor_cost', 0)
            discount_text = f"İndirim (%{discount_percentage}): <font color='#dc2626'>-₺ {self._format_price_modern(discount_amount)}</font>"
            totals_content.append(Paragraph(discount_text, self.normal_style))
            totals_content.append(Spacer(1, 4))
        
        # İşçilik maliyeti (eğer varsa)
        labor_cost = quote_data.get('labor_cost', 0)
        if labor_cost > 0:
            labor_text = f"İşçilik: <font color='#059669'>+₺ {self._format_price_modern(labor_cost)}</font>"
            totals_content.append(Paragraph(labor_text, self.normal_style))
            totals_content.append(Spacer(1, 8))
        
        # Net toplam - küçültülmüş ve yeni renk (#2F4B68)
        net_total = quote_data.get('total_net_price', 0)
        net_total_color = colors.HexColor('#2F4B68')  # YENİ RENK
        net_total_text = f"<font size='14' color='{net_total_color}'><b>NET TOPLAM: ₺ {self._format_price_modern(net_total)}</b></font>"
        totals_content.append(Paragraph(net_total_text, self.price_style))
        
        return totals_content
    
    def _create_modern_footer(self):
        """Güncellenmiş footer mesajı - Yeni renk ile"""
        footer_content = []
        
        # Önemli notlar başlığı - YENİ RENK (#2F4B68)
        footer_content.append(Paragraph("<b><font color='#2F4B68'>ÖNEMLİ NOTLAR:</font></b>", self.subtitle_style))
        footer_content.append(Spacer(1, 8))
        
        # Güncellenmiş notlar listesi
        notes = [
            "• Yukarıdaki fiyatlandırmanın, 1 hafta geçerli olduğunu lütfen göz önünde bulundurunuz.",
            "• Ürün özellikleri ve fiyatları değişiklik gösterebilir.",
            "• Servisimiz dışında yapılan işlemlerde montaj ve nakliye masrafları ayrıca hesaplanacaktır."
        ]
        
        notes_text = "<br/>".join(notes)
        footer_content.append(Paragraph(notes_text, self.footer_style))
        
        return footer_content
    
    def _format_price_modern(self, price):
        """Modern Türkçe fiyat formatla"""
        try:
            if price is None:
                return "0,00"
            
            price_float = float(price)
            if price_float == 0:
                return "0,00"
                
            # Türkçe format: nokta binlik ayırıcı, virgül ondalık ayırıcı
            formatted = f"{price_float:,.2f}"
            # Binlik ayırıcıyı nokta, ondalık ayırıcıyı virgül yap
            formatted = formatted.replace(',', 'TEMP').replace('.', ',').replace('TEMP', '.')
            return formatted
        except (ValueError, TypeError):
            return "0,00"
    
    def _format_price(self, price):
        """Eski format - geriye uyumluluk için"""
        return self._format_price_modern(price)

# ===== PDF QUOTE ENDPOINT =====
# ===== PACKAGE PDF GENERATOR =====

class PDFPackageGenerator(PDFQuoteGenerator):
    """Paket PDF oluşturucu - Teklif taslağını kullanan"""
    
    def __init__(self):
        super().__init__()  # PDFQuoteGenerator'dan miras al
        # Eksik style'ları ekle
        self.header_style = ParagraphStyle(
            'PackageHeader',
            parent=self.styles['Normal'],
            fontName=self.get_font_name(is_bold=True),
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.white
        )

    def _format_price_modern(self, price):
        """Modern format ile fiyat gösterimi"""
        if price is None or price == 0:
            return "0,00"
        
        # Format price with Turkish decimal comma
        formatted = f"{float(price):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return formatted

    def generate_package_pdf(self, package_data, products, include_prices=True, categories=None, category_groups=None):
        """Teklif taslağını kullanarak paket PDF'i oluştur"""
        buffer = BytesIO()
        
        # Veriler parametre olarak geçildiyse onları kullan, yoksa boş liste
        if categories is None:
            categories = []
        if category_groups is None:
            category_groups = []
        
        # Yüksek kaliteli PDF ayarları (teklif ile aynı)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2.5*cm,
            leftMargin=2.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
            title=f"Paket - {package_data.get('name', 'Adsız')}"
        )
        
        # PDF içeriği
        story = []
        
        # Header (Logo + Firma bilgileri)
        story.append(self._create_modern_header())
        story.append(Spacer(1, 30))
        
        # Tarih - sağ üst köşe
        story.append(self._create_date_header())
        story.append(Spacer(1, 10))
        
        # Paket başlığı
        package_name = package_data.get('name', 'Paket Bilgisi')
        story.append(Paragraph(f"<b>{package_name}</b>", self.title_style))
        story.append(Spacer(1, 25))
        
        # Ürün tablosu başlığı
        story.append(Paragraph("<b>Paket İçeriği</b>", self.subtitle_style))
        story.append(Spacer(1, 10))
        
        # Ürün tablosu (kategori grupları ile)
        story.append(self._create_package_products_table_with_groups(products, include_prices, categories, category_groups))
        story.append(Spacer(1, 25))
        
        # Toplam hesaplama bölümü (indirim ve işçilik ile)
        if include_prices:
            # Fiyatlı listede indirim ve işçilik hesaplaması
            total_list_price = sum(float(p.get('list_price_try', 0)) * p.get('quantity', 1) for p in products)
            discount_percentage = float(package_data.get('discount_percentage', 0))
            labor_cost = float(package_data.get('labor_cost', 0))
            
            discount_amount = total_list_price * (discount_percentage / 100)
            total_after_discount = total_list_price - discount_amount
            final_total = total_after_discount + labor_cost
            
            story.extend(self._create_package_totals_section_with_discount_labor(
                total_list_price, discount_percentage, discount_amount, labor_cost, final_total
            ))
        else:
            # Fiyatsız listede sadece satış fiyatı
            sale_price = float(package_data.get('sale_price', 0))
            story.extend(self._create_package_totals_section(sale_price, "Paket Satış Fiyatı"))
        
        story.append(Spacer(1, 30))
        
        # Footer notları
        story.extend(self._create_modern_footer())
        
        # PDF oluştur
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    # _create_package_info_section removed - date moved to top-right
    
    def _create_package_products_table(self, products, include_prices=True):
        """Paket ürünleri tablosu - teklif stilinde"""
        from reportlab.platypus import Table as PDFTable
        
        # Tablo başlıkları
        if include_prices:
            headers = ["Ürün Adı", "Adet", "Birim Fiyat", "Toplam"]
            col_widths = [8*cm, 2*cm, 3*cm, 3*cm]
        else:
            headers = ["Ürün Adı", "Adet"]
            col_widths = [12*cm, 3*cm]
        
        # Header row
        header_row = []
        for header in headers:
            header_row.append(Paragraph(f"<b>{header}</b>", self.header_style))
        
        table_data = [header_row]
        
        # Ürün satırları
        for product in products:
            quantity = product.get('quantity', 1)
            product_name = product.get('name', '')
            
            if include_prices:
                unit_price = float(product.get('list_price_try', 0))
                line_total = unit_price * quantity
                
                row = [
                    Paragraph(product_name, self.data_style),
                    Paragraph(str(quantity), self.data_style),
                    Paragraph(f"₺ {self._format_price_modern(unit_price)}", self.data_style),
                    Paragraph(f"₺ {self._format_price_modern(line_total)}", self.data_style)
                ]
            else:
                row = [
                    Paragraph(product_name, self.data_style),
                    Paragraph(str(quantity), self.data_style)
                ]
            
            table_data.append(row)
        
        # Tablo oluştur
        table = PDFTable(table_data, colWidths=col_widths, repeatRows=1)
        
        # Tablo stili (teklif ile aynı)
        table.setStyle(TableStyle([
            # Header stili
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F4B68')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.get_font_name(is_bold=True)),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Veri stili
            ('FONTNAME', (0, 1), (-1, -1), self.get_font_name()),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Ürün adları sola hizalı
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Diğer kolonlar orta hizalı
            ('LEFTPADDING', (0, 1), (0, -1), 8),
            ('RIGHTPADDING', (-1, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            
            # Kenarlık ve zemin
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')])
        ]))
        
        return table
    
    def _create_package_products_table_with_groups(self, products, include_prices=True, categories=None, category_groups=None):
        """Paket ürünleri tablosu - kategori grupları ile organize edilmiş"""
        from reportlab.platypus import Table as PDFTable
        
        # Kategorileri ve kategori gruplarını kullan (parametre olarak geçildi)
        if categories is None:
            categories = []
        if category_groups is None:
            category_groups = []
        
        # Kategori grup haritası oluştur
        category_to_group = {}
        for group in category_groups:
            for cat_id in group.get("category_ids", []):
                category_to_group[cat_id] = group.get("name", "")
        
        # Ürünleri kategori gruplarına göre grupla
        grouped_products = {}
        for product in products:
            category_id = product.get('category_id')
            category = next((c for c in categories if c.get('id') == category_id), None)
            
            # Kategori grubunu belirle
            if category_id and category_id in category_to_group:
                group_name = category_to_group[category_id]
            elif category:
                group_name = category.get('name', 'Diğer')
            else:
                group_name = 'Kategorisiz'
            
            if group_name not in grouped_products:
                grouped_products[group_name] = []
            grouped_products[group_name].append(product)
        
        # Tablo başlıkları
        if include_prices:
            headers = ["Ürün Adı", "Adet", "Birim Fiyat", "Toplam"]
            col_widths = [8*cm, 2*cm, 3*cm, 3*cm]
        else:
            headers = ["Ürün Adı", "Adet"]
            col_widths = [12*cm, 3*cm]
        
        # Header row
        header_row = []
        for header in headers:
            header_row.append(Paragraph(f"<b>{header}</b>", self.header_style))
        
        table_data = [header_row]
        
        # Grup başlıklarını ve ürünleri ekle
        for group_name, group_products in grouped_products.items():
            # Grup başlığı satırı - küçültülmüş font
            group_header_style = ParagraphStyle(
                'GroupHeader',
                parent=self.styles['Normal'],
                fontName=self.get_font_name(is_bold=True),
                fontSize=7,  # Daha küçük font
                alignment=TA_LEFT,
                textColor=colors.HexColor('#2F4B68'),
                leftIndent=5
            )
            
            if include_prices:
                group_row = [
                    Paragraph(f"<b>{group_name}</b>", group_header_style),
                    Paragraph("", group_header_style),
                    Paragraph("", group_header_style),
                    Paragraph("", group_header_style)
                ]
            else:
                group_row = [
                    Paragraph(f"<b>{group_name}</b>", group_header_style),
                    Paragraph("", group_header_style)
                ]
            
            table_data.append(group_row)
            
            # Grup içindeki ürünler - çok küçültülmüş font
            small_data_style = ParagraphStyle(
                'SmallData',
                parent=self.styles['Normal'],
                fontName=self.get_font_name(),
                fontSize=6,  # Yarıdan da küçük (9'dan 6'ya)
                alignment=TA_LEFT,
                leftIndent=10
            )
            
            for product in group_products:
                quantity = product.get('quantity', 1)
                product_name = product.get('name', '')
                
                if include_prices:
                    unit_price = float(product.get('list_price_try', 0))
                    line_total = unit_price * quantity
                    
                    row = [
                        Paragraph(f"• {product_name}", small_data_style),
                        Paragraph(str(quantity), small_data_style),
                        Paragraph(f"₺ {self._format_price_modern(unit_price)}", small_data_style),
                        Paragraph(f"₺ {self._format_price_modern(line_total)}", small_data_style)
                    ]
                else:
                    row = [
                        Paragraph(f"• {product_name}", small_data_style),
                        Paragraph(str(quantity), small_data_style)
                    ]
                
                table_data.append(row)
        
        # Tablo oluştur
        table = PDFTable(table_data, colWidths=col_widths, repeatRows=1)
        
        # Tablo stili
        styles_list = [
            # Header stili
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F4B68')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.get_font_name(is_bold=True)),
            ('FONTSIZE', (0, 0), (-1, 0), 8),  # Header da küçültüldü
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),  # Padding küçültüldü
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Veri stili
            ('FONTNAME', (0, 1), (-1, -1), self.get_font_name()),
            ('FONTSIZE', (0, 1), (-1, -1), 6),  # Çok küçültüldü
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Ürün adları sola hizalı
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),  # Diğer kolonlar orta hizalı
            ('LEFTPADDING', (0, 1), (0, -1), 4),  # Padding küçültüldü
            ('RIGHTPADDING', (-1, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            
            # Kenarlık ve zemin
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),  # İnce kenarlık
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
        ]
        
        # Grup başlıkları için özel stil ekle
        row_idx = 1  # Header'dan sonra
        for group_name, group_products in grouped_products.items():
            # Grup başlığı satırı için özel background
            styles_list.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#EDF2F7')))
            styles_list.append(('FONTSIZE', (0, row_idx), (-1, row_idx), 7))
            row_idx += 1  # Grup başlığı
            row_idx += len(group_products)  # Grup ürünleri
        
        table.setStyle(TableStyle(styles_list))
        
        return table

    def _create_package_totals_section_with_discount_labor(self, total_list_price, discount_percentage, discount_amount, labor_cost, final_total):
        """Create package totals section with discount and labor cost calculations"""
        from reportlab.platypus import Table as PDFTable
        
        elements = []
        
        # Başlık
        elements.append(Paragraph("<b>Paket Fiyat Hesaplaması</b>", self.subtitle_style))
        elements.append(Spacer(1, 10))
        
        # Tablo verisi
        table_data = [
            [Paragraph("<b>Açıklama</b>", self.header_style), Paragraph("<b>Tutar</b>", self.header_style)]
        ]
        
        # Liste fiyatı toplamı
        table_data.append([
            Paragraph("Toplam Liste Fiyatı", self.data_style),
            Paragraph(f"₺ {self._format_price_modern(total_list_price)}", self.data_style)
        ])
        
        # İndirim (eğer varsa)
        if discount_percentage > 0:
            table_data.append([
                Paragraph(f"İndirim (%{discount_percentage})", self.data_style),
                Paragraph(f"- ₺ {self._format_price_modern(discount_amount)}", 
                         ParagraphStyle('RedText', parent=self.data_style, textColor=colors.red))
            ])
            
            table_data.append([
                Paragraph("İndirim Sonrası Toplam", self.data_style),
                Paragraph(f"₺ {self._format_price_modern(total_list_price - discount_amount)}", self.data_style)
            ])
        
        # İşçilik maliyeti (eğer varsa)
        if labor_cost > 0:
            table_data.append([
                Paragraph("İşçilik Maliyeti", self.data_style),
                Paragraph(f"+ ₺ {self._format_price_modern(labor_cost)}", 
                         ParagraphStyle('GreenText', parent=self.data_style, textColor=colors.green))
            ])
        
        # Net toplam
        table_data.append([
            Paragraph("<b>Net Toplam</b>", 
                     ParagraphStyle('BoldData', parent=self.data_style, fontName=self.get_font_name(is_bold=True))),
            Paragraph(f"<b>₺ {self._format_price_modern(final_total)}</b>", 
                     ParagraphStyle('BoldTotal', parent=self.data_style, 
                                   fontName=self.get_font_name(is_bold=True), 
                                   textColor=colors.HexColor('#2F4B68')))
        ])
        
        # Tablo oluştur
        table = PDFTable(table_data, colWidths=[10*cm, 6*cm])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F4B68')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.get_font_name(is_bold=True)),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Veri satırları
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), self.get_font_name()),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('LEFTPADDING', (0, 1), (-1, -1), 10),
            ('RIGHTPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            
            # Kenarlık
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#2F4B68')),  # Net toplam üstüne kalın çizgi
            
            # Zemin renkleri
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EDF2F7'))  # Net toplam arka planı
        ]))
        
        elements.append(table)
        return elements
    
    def _create_package_totals_section(self, amount, label):
        """Paket toplam bölümü"""
        from reportlab.platypus import Table as PDFTable
        
        # Toplam tablosu
        totals_data = [
            [label, f"₺ {self._format_price_modern(amount)}"]
        ]
        
        totals_table = PDFTable(totals_data, colWidths=[10*cm, 6*cm])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'RIGHT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), self.get_font_name(is_bold=True)),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2F4B68')),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#2F4B68')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        return [totals_table]

# ===== PACKAGE PDF ENDPOINTS =====

@app.get("/api/packages/{package_id}/pdf-with-prices")
async def download_package_pdf_with_prices(package_id: str):
    """Paket PDF'i indir - ürün isimleri ve liste fiyatları ile"""
    try:
        # Paket ve ürünleri getir
        package = await db.packages.find_one({"id": package_id})
        if not package:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        # Paket ürünlerini getir
        package_products = await db.package_products.find({"package_id": package_id}).to_list(None)
        
        # Ürün detaylarını al
        products = []
        for pp in package_products:
            product = await db.products.find_one({"id": pp["product_id"]})
            if product:
                product_data = {
                    "name": product["name"],
                    "quantity": pp["quantity"],
                    "list_price_try": product.get("list_price_try", 0),
                    "category_id": product.get("category_id")  # Kategori bilgisi eklendi
                }
                products.append(product_data)
        
        # Kategorileri ve kategori gruplarını önceden getir
        categories = await db.categories.find().to_list(None)
        category_groups = await db.category_groups.find().to_list(None)
        
        # PDF oluştur
        generator = PDFPackageGenerator()
        pdf_buffer = generator.generate_package_pdf(package, products, include_prices=True, categories=categories, category_groups=category_groups)
        
        # Dosya adı
        safe_name = "".join(c for c in package["name"] if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"paket_{safe_name}_fiyatli.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating package PDF with prices: {e}")
        raise HTTPException(status_code=500, detail="PDF oluşturulamadı")

@app.get("/api/packages/{package_id}/pdf-without-prices")
async def download_package_pdf_without_prices(package_id: str):
    """Paket PDF'i indir - sadece ürün isimleri ile"""
    try:
        # Paket ve ürünleri getir
        package = await db.packages.find_one({"id": package_id})
        if not package:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        # Paket ürünlerini getir
        package_products = await db.package_products.find({"package_id": package_id}).to_list(None)
        
        # Ürün detaylarını al
        products = []
        for pp in package_products:
            product = await db.products.find_one({"id": pp["product_id"]})
            if product:
                product_data = {
                    "name": product["name"],
                    "quantity": pp["quantity"],
                    "category_id": product.get("category_id")  # Kategori bilgisi eklendi
                }
                products.append(product_data)
        
        # Kategorileri ve kategori gruplarını önceden getir
        categories = await db.categories.find().to_list(None)
        category_groups = await db.category_groups.find().to_list(None)
        
        # PDF oluştur
        generator = PDFPackageGenerator()
        pdf_buffer = generator.generate_package_pdf(package, products, include_prices=False, categories=categories, category_groups=category_groups)
        
        # Dosya adı
        safe_name = "".join(c for c in package["name"] if c.isalnum() or c in (' ', '-', '_')).strip()
        filename = f"paket_{safe_name}_liste.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating package PDF without prices: {e}")
        raise HTTPException(status_code=500, detail="PDF oluşturulamadı")


@app.get("/api/quotes/{quote_id}/pdf")
async def download_quote_pdf(quote_id: str):
    """Teklif PDF'ini indir"""
    try:
        db = await get_db()
        quote = await db.quotes.find_one({"id": quote_id})
        
        if not quote:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        # PDF oluştur
        pdf_generator = PDFQuoteGenerator()
        pdf_buffer = pdf_generator.create_quote_pdf(quote)
        
        # Response headers - ensure proper encoding for Turkish characters
        safe_filename = quote["name"].encode('ascii', 'ignore').decode('ascii')
        if not safe_filename:
            safe_filename = "teklif"
        
        headers = {
            'Content-Type': 'application/pdf',
            'Content-Disposition': f'attachment; filename="{safe_filename}.pdf"'
        }
        
        return StreamingResponse(
            BytesIO(pdf_buffer.read()),
            media_type='application/pdf',
            headers=headers
        )
        
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")

# Category endpoints
@api_router.post("/categories", response_model=Category)
async def create_category(category: CategoryCreate):
    """Create a new category"""
    try:
        category_dict = {
            "id": str(uuid.uuid4()),
            "name": category.name,
            "description": category.description,
            "color": category.color or "#3B82F6",  # Default blue color
            "created_at": datetime.now(timezone.utc)
        }
        
        result = await db.categories.insert_one(category_dict)
        return Category(**category_dict)
        
    except Exception as e:
        logger.error(f"Error creating category: {e}")
        raise HTTPException(status_code=500, detail="Kategori oluşturulamadı")

@api_router.get("/categories", response_model=List[Category])
async def get_categories():
    """Get all categories sorted by sort_order, then by name"""
    try:
        # Kategorileri önce sort_order'a, sonra name'e göre sırala
        categories = await db.categories.find().sort([("sort_order", 1), ("name", 1)]).to_list(None)
        return [Category(**category) for category in categories]
    except Exception as e:
        logger.error(f"Error getting categories: {e}")
        raise HTTPException(status_code=500, detail="Kategoriler getirilemedi")

@api_router.patch("/categories/{category_id}")
async def update_category(category_id: str, update_data: CategoryCreate):
    """Update a category"""
    try:
        update_dict = {}
        if update_data.name:
            update_dict["name"] = update_data.name
        if update_data.description is not None:
            update_dict["description"] = update_data.description
        if update_data.color:
            update_dict["color"] = update_data.color
        if update_data.sort_order is not None:
            update_dict["sort_order"] = update_data.sort_order
        
        if update_dict:
            result = await db.categories.update_one(
                {"id": category_id},
                {"$set": update_dict}
            )
            
            if result.modified_count == 0:
                raise HTTPException(status_code=404, detail="Kategori bulunamadı")
        
        # Get updated category
        updated_category = await db.categories.find_one({"id": category_id})
        return {
            "success": True,
            "message": "Kategori başarıyla güncellendi",
            "category": Category(**updated_category) if updated_category else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating category: {e}")
        raise HTTPException(status_code=500, detail="Kategori güncellenemedi")

@api_router.post("/categories/reorder")
async def reorder_categories(category_orders: List[Dict[str, Any]]):
    """Kategorilerin sırasını toplu güncelle"""
    try:
        # category_orders: [{"id": "cat1", "sort_order": 1}, {"id": "cat2", "sort_order": 2}, ...]
        for item in category_orders:
            category_id = item.get("id")
            sort_order = item.get("sort_order", 0)
            
            if category_id:
                await db.categories.update_one(
                    {"id": category_id},
                    {"$set": {"sort_order": sort_order}}
                )
        
        # Return updated categories sorted by new order
        categories = await db.categories.find().sort([("sort_order", 1), ("name", 1)]).to_list(None)
        return {
            "success": True,
            "message": "Kategori sıralaması güncellendi",
            "categories": [Category(**category) for category in categories]
        }
        
    except Exception as e:
        logger.error(f"Error reordering categories: {e}")
        raise HTTPException(status_code=500, detail="Kategori sıralaması güncellenemedi")

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str):
    """Delete a category"""
    try:
        # Check if category exists and is deletable
        category = await db.categories.find_one({"id": category_id})
        if not category:
            raise HTTPException(status_code=404, detail="Kategori bulunamadı")
        
        # Check if category is deletable
        if not category.get("is_deletable", True):
            raise HTTPException(status_code=400, detail="Bu kategori silinemez")
        
        # First, remove this category from all products
        await db.products.update_many(
            {"category_id": category_id},
            {"$unset": {"category_id": ""}}
        )
        
        # Then delete the category
        result = await db.categories.delete_one({"id": category_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Kategori bulunamadı")
        
        return {"success": True, "message": "Kategori başarıyla silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting category: {e}")
        raise HTTPException(status_code=500, detail="Kategori silinemedi")

@api_router.post("/packages/{package_id}/supplies")
async def add_supplies_to_package(package_id: str, supplies: List[PackageSupplyCreate]):
    """Pakete sarf malzemesi ekle"""
    try:
        # Check if package exists
        package = await db.packages.find_one({"id": package_id})
        if not package:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        # Remove existing supplies
        await db.package_supplies.delete_many({"package_id": package_id})
        
        # Add new supplies
        package_supplies = []
        for supply in supplies:
            # Verify product exists
            existing_product = await db.products.find_one({"id": supply.product_id})
            if not existing_product:
                continue
                
            package_supply = {
                "id": str(uuid.uuid4()),
                "package_id": package_id,
                "product_id": supply.product_id,
                "quantity": supply.quantity,
                "note": supply.note,
                "created_at": datetime.now(timezone.utc)
            }
            package_supplies.append(package_supply)
        
        if package_supplies:
            await db.package_supplies.insert_many(package_supplies)
        
        return {"success": True, "message": f"{len(package_supplies)} sarf malzemesi pakete eklendi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding supplies to package: {e}")
        raise HTTPException(status_code=500, detail="Sarf malzemeleri pakete eklenemedi")

@api_router.delete("/packages/{package_id}/supplies/{supply_id}")
async def remove_supply_from_package(package_id: str, supply_id: str):
    """Paketten sarf malzemesi çıkar"""
    try:
        # supply_id actually refers to product_id in this context
        result = await db.package_supplies.delete_one({
            "product_id": supply_id,  # Changed from "id" to "product_id"
            "package_id": package_id
        })
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Sarf malzemesi bulunamadı")
        
        return {"success": True, "message": "Sarf malzemesi paketten çıkarıldı"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error removing supply from package: {e}")
        raise HTTPException(status_code=500, detail="Sarf malzemesi paketten çıkarılamadı")

@api_router.put("/packages/{package_id}/supplies/{supply_id}")
async def update_supply_quantity(package_id: str, supply_id: str, quantity: int):
    """Sarf malzemesi adetini güncelle"""
    try:
        if quantity <= 0:
            raise HTTPException(status_code=400, detail="Adet 1'den küçük olamaz")
        
        # supply_id actually refers to product_id in this context
        result = await db.package_supplies.update_one(
            {"product_id": supply_id, "package_id": package_id},  # Changed from "id" to "product_id"
            {"$set": {"quantity": quantity}}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Sarf malzemesi bulunamadı")
        
        return {"success": True, "message": "Sarf malzemesi adeti güncellendi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating supply quantity: {e}")
        raise HTTPException(status_code=500, detail="Sarf malzemesi adeti güncellenemedi")
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Kategori bulunamadı")
        
        return {"success": True, "message": "Kategori silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting category: {e}")
        raise HTTPException(status_code=500, detail="Kategori silinemedi")

@api_router.get("/products/supplies")
async def get_supply_products():
    """Get products from 'Sarf Malzemeleri' category only"""
    try:
        # Get Sarf Malzemeleri category
        supplies_category = await db.categories.find_one({"name": "Sarf Malzemeleri"})
        if not supplies_category:
            return []
        
        # Get products from supplies category
        products = await db.products.find({
            "category_id": supplies_category["id"]
        }).sort("name", 1).to_list(None)
        
        return [Product(**product) for product in products]
    except Exception as e:
        logger.error(f"Error getting supply products: {e}")
        raise HTTPException(status_code=500, detail="Sarf malzemesi ürünleri getirilemedi")
        raise
    except Exception as e:
        logger.error(f"Error deleting category: {e}")
        raise HTTPException(status_code=500, detail="Kategori silinemedi")

@api_router.post("/products/{product_id}/assign-category")
async def assign_product_to_category(product_id: str, category_id: str = None):
    """Assign a product to a category"""
    try:
        update_dict = {"category_id": category_id} if category_id else {"$unset": {"category_id": ""}}
        
        result = await db.products.update_one(
            {"id": product_id},
            update_dict if category_id else {"$unset": {"category_id": ""}}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        return {"success": True, "message": "Ürün kategoriye atandı"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error assigning product to category: {e}")
        raise HTTPException(status_code=500, detail="Ürün kategoriye atanamadı")

@api_router.post("/products/{product_id}/toggle-favorite")
async def toggle_product_favorite(product_id: str):
    """Toggle a product's favorite status"""
    try:
        # Get current product
        current_product = await db.products.find_one({"id": product_id})
        if not current_product:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        # Toggle favorite status
        current_favorite = current_product.get("is_favorite", False)
        new_favorite = not current_favorite
        
        result = await db.products.update_one(
            {"id": product_id},
            {"$set": {"is_favorite": new_favorite}}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Ürün bulunamadı")
        
        status_message = "favorilere eklendi" if new_favorite else "favorilerden çıkarıldı"
        return {
            "success": True, 
            "message": f"Ürün {status_message}",
            "is_favorite": new_favorite
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error toggling product favorite: {e}")
        raise HTTPException(status_code=500, detail="Favori durumu güncellenemedi")

@api_router.post("/products/{product_id}/favorite")
async def toggle_product_favorite_v2(product_id: str):
    """Toggle favorite status of a product"""
    try:
        # Get current product
        product = await db.products.find_one({"id": product_id})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Toggle favorite status
        new_favorite_status = not product.get("is_favorite", False)
        
        update_data = {"is_favorite": new_favorite_status}
        
        # If removing from favorites, clear stock quantity
        if not new_favorite_status:
            update_data["stock_quantity"] = None
        
        await db.products.update_one(
            {"id": product_id},
            {"$set": update_data}
        )
        
        return {"success": True, "is_favorite": new_favorite_status}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error toggling product favorite: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@api_router.post("/products/{product_id}/stock")
async def update_product_stock(product_id: str, stock_quantity: int = Form(...)):
    """Update stock quantity for a favorite product"""
    try:
        # Get current product
        product = await db.products.find_one({"id": product_id})
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Check if product is favorite
        if not product.get("is_favorite", False):
            raise HTTPException(status_code=400, detail="Stok sadece favori ürünler için takip edilir")
        
        # Update stock quantity
        await db.products.update_one(
            {"id": product_id},
            {"$set": {"stock_quantity": stock_quantity}}
        )
        
        return {
            "success": True, 
            "message": "Stok miktarı güncellendi",
            "stock_quantity": stock_quantity
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating product stock: {e}")
        raise HTTPException(status_code=500, detail="Stok güncellenirken hata oluştu")
@api_router.get("/products/favorites", response_model=List[Product])
async def get_favorite_products():
    """Get all favorite products"""
    try:
        products = await db.products.find({"is_favorite": True}).sort("name", 1).to_list(None)
        return [Product(**product) for product in products]
    except Exception as e:
        logger.error(f"Error getting favorite products: {e}")
        raise HTTPException(status_code=500, detail="Favori ürünler getirilemedi")

# Package endpoints
@api_router.post("/packages", response_model=Package)
async def create_package(package: PackageCreate):
    """Create a new package"""
    try:
        package_data = package.dict()
        package_data["id"] = str(uuid.uuid4())
        package_data["created_at"] = datetime.now(timezone.utc)
        
        # Convert Decimal to float for MongoDB compatibility
        if "sale_price" in package_data and package_data["sale_price"] is not None:
            package_data["sale_price"] = float(package_data["sale_price"])
        
        result = await db.packages.insert_one(package_data)
        if result.inserted_id:
            created_package = await db.packages.find_one({"id": package_data["id"]})
            return Package(**created_package)
        else:
            raise HTTPException(status_code=500, detail="Paket oluşturulamadı")
    except Exception as e:
        logger.error(f"Error creating package: {e}")
        raise HTTPException(status_code=500, detail="Paket oluşturulamadı")

@api_router.get("/packages", response_model=List[Package])
async def get_packages():
    """Get all packages sorted with pinned packages first"""
    try:
        # Get packages sorted by pin status (pinned first) then by creation date (newest first)
        packages = await db.packages.find({}).sort([
            ("is_pinned", -1),  # Pinned packages first (True = -1 comes before False = 0)
            ("created_at", -1)   # Then by creation date, newest first
        ]).to_list(None)
        
        return [Package(**package) for package in packages]
    except Exception as e:
        logger.error(f"Error getting packages: {e}")
        raise HTTPException(status_code=500, detail="Paketler getirilemedi")

@api_router.get("/packages/{package_id}", response_model=PackageWithProducts)
async def get_package_with_products(package_id: str):
    """Get package with its products"""
    try:
        # Get package
        package = await db.packages.find_one({"id": package_id})
        if not package:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        # Get package products
        package_products = await db.package_products.find({"package_id": package_id}).to_list(None)
        
        # Get product details and calculate totals
        products = []
        total_discounted_price = Decimal('0')
        
        for pp in package_products:
            product = await db.products.find_one({"id": pp["product_id"]})
            if product:
                product_data = {
                    "id": product["id"],
                    "name": product["name"],
                    "list_price": product.get("list_price", 0),
                    "discounted_price": product.get("discounted_price"),
                    "list_price_try": product.get("list_price_try", 0),
                    "discounted_price_try": product.get("discounted_price_try"),
                    "currency": product.get("currency", "USD"),
                    "quantity": pp["quantity"],
                    "company_id": product.get("company_id"),
                    "category_id": product.get("category_id")
                }
                products.append(product_data)
                
                # Calculate discounted price total
                if product.get("discounted_price_try"):
                    total_discounted_price += Decimal(str(product["discounted_price_try"])) * pp["quantity"]
                else:
                    total_discounted_price += Decimal(str(product.get("list_price_try", 0))) * pp["quantity"]

        # Get package supplies (sarf malzemeleri)
        package_supplies = await db.package_supplies.find({"package_id": package_id}).to_list(None)
        
        # Get supply details and calculate totals
        supplies = []
        total_supplies_price = Decimal('0')
        
        for ps in package_supplies:
            supply = await db.products.find_one({"id": ps["product_id"]})
            if supply:
                supply_data = {
                    "id": supply["id"],
                    "name": supply["name"],
                    "list_price": supply.get("list_price", 0),
                    "discounted_price": supply.get("discounted_price"),
                    "list_price_try": supply.get("list_price_try", 0),
                    "discounted_price_try": supply.get("discounted_price_try"),
                    "currency": supply.get("currency", "USD"),
                    "quantity": ps["quantity"],
                    "note": ps.get("note", ""),
                    "company_id": supply.get("company_id"),
                    "category_id": supply.get("category_id")
                }
                supplies.append(supply_data)
                
                # Calculate supply price total
                if supply.get("discounted_price_try"):
                    total_supplies_price += Decimal(str(supply["discounted_price_try"])) * ps["quantity"]
                else:
                    total_supplies_price += Decimal(str(supply.get("list_price_try", 0))) * ps["quantity"]
        
        return PackageWithProducts(
            id=package["id"],
            name=package["name"],
            description=package.get("description"),
            sale_price=package["sale_price"],
            discount_percentage=package.get("discount_percentage", 0),  # İndirim yüzdesi eklendi
            labor_cost=package.get("labor_cost", 0),  # İşçilik maliyeti eklendi
            image_url=package.get("image_url"),
            created_at=package["created_at"],
            products=products,
            supplies=supplies,
            total_discounted_price=total_discounted_price,
            total_discounted_price_with_supplies=total_discounted_price + total_supplies_price,
            status="active"  # Default status
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting package with products: {e}")
        raise HTTPException(status_code=500, detail="Paket detayları getirilemedi")

@api_router.put("/packages/{package_id}", response_model=Package)
async def update_package(package_id: str, package: PackageCreate):
    """Update a package"""
    try:
        update_data = package.dict(exclude_unset=True)
        
        # Convert Decimal to float for MongoDB compatibility
        if "sale_price" in update_data and update_data["sale_price"] is not None:
            update_data["sale_price"] = float(update_data["sale_price"])
        
        result = await db.packages.update_one(
            {"id": package_id},
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        updated_package = await db.packages.find_one({"id": package_id})
        return Package(**updated_package)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating package: {e}")
        raise HTTPException(status_code=500, detail="Paket güncellenemedi")

@api_router.post("/packages/{package_id}/pin")
async def toggle_package_pin(package_id: str):
    """Toggle pin status of a package"""
    try:
        # Get current package
        package = await db.packages.find_one({"id": package_id})
        if not package:
            raise HTTPException(status_code=404, detail="Package not found")
        
        # Toggle pin status
        new_pin_status = not package.get("is_pinned", False)
        
        await db.packages.update_one(
            {"id": package_id},
            {"$set": {"is_pinned": new_pin_status}}
        )
        
        action = "sabitlendi" if new_pin_status else "sabitleme kaldırıldı"
        
        return {
            "success": True,
            "message": f"Paket başarıyla {action}",
            "is_pinned": new_pin_status,
            "package_name": package["name"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error toggling package pin: {e}")
        raise HTTPException(status_code=500, detail="Paket sabitleme durumu değiştirilemedi")

@api_router.post("/packages/{package_id}/copy")
async def copy_package(package_id: str, new_name: str = Form(...)):
    """Copy a package with all its products and supplies"""
    try:
        # Get original package
        original_package = await db.packages.find_one({"id": package_id})
        if not original_package:
            raise HTTPException(status_code=404, detail="Package not found")
        
        # Check if new name already exists
        existing_package = await db.packages.find_one({"name": new_name})
        if existing_package:
            raise HTTPException(status_code=400, detail="Bu isimde bir paket zaten mevcut")
        
        # Create new package with copied data
        new_package_id = str(uuid.uuid4())
        new_package = {
            "id": new_package_id,
            "name": new_name,
            "description": original_package.get("description", ""),
            "sale_price": original_package.get("sale_price", "0"),
            "image_url": original_package.get("image_url"),
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.packages.insert_one(new_package)
        logger.info(f"Created package copy: {new_name} (ID: {new_package_id})")
        
        # Copy package products
        original_products = await db.package_products.find({"package_id": package_id}).to_list(length=None)
        if original_products:
            copied_products = []
            for product in original_products:
                copied_product = {
                    "id": str(uuid.uuid4()),
                    "package_id": new_package_id,
                    "product_id": product["product_id"],
                    "quantity": product["quantity"],
                    "created_at": datetime.now(timezone.utc)
                }
                copied_products.append(copied_product)
            
            await db.package_products.insert_many(copied_products)
            logger.info(f"Copied {len(copied_products)} products to new package")
        
        # Copy package supplies
        original_supplies = await db.package_supplies.find({"package_id": package_id}).to_list(length=None)
        if original_supplies:
            copied_supplies = []
            for supply in original_supplies:
                copied_supply = {
                    "id": str(uuid.uuid4()),
                    "package_id": new_package_id,
                    "product_id": supply["product_id"],
                    "quantity": supply["quantity"],
                    "created_at": datetime.now(timezone.utc)
                }
                copied_supplies.append(copied_supply)
            
            await db.package_supplies.insert_many(copied_supplies)
            logger.info(f"Copied {len(copied_supplies)} supplies to new package")
        
        return {
            "success": True, 
            "message": f"Paket başarıyla kopyalandı: {new_name}",
            "new_package_id": new_package_id,
            "original_package_name": original_package["name"],
            "new_package_name": new_name,
            "copied_products": len(original_products) if original_products else 0,
            "copied_supplies": len(original_supplies) if original_supplies else 0
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error copying package: {e}")
        raise HTTPException(status_code=500, detail="Paket kopyalanırken hata oluştu")

@api_router.delete("/packages/{package_id}")
async def delete_package(package_id: str):
    """Delete a package and its products"""
    try:
        # Delete package products and supplies first
        await db.package_products.delete_many({"package_id": package_id})
        await db.package_supplies.delete_many({"package_id": package_id})
        
        # Delete package
        result = await db.packages.delete_one({"id": package_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        return {"success": True, "message": "Paket başarıyla silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting package: {e}")
        raise HTTPException(status_code=500, detail="Paket silinemedi")

@api_router.post("/packages/{package_id}/products")
async def add_products_to_package(package_id: str, products: List[PackageProductCreate]):
    """Add products to a package"""
    try:
        # Check if package exists
        package = await db.packages.find_one({"id": package_id})
        if not package:
            raise HTTPException(status_code=404, detail="Paket bulunamadı")
        
        # Remove existing products
        await db.package_products.delete_many({"package_id": package_id})
        
        # Add new products
        package_products = []
        for product in products:
            # Verify product exists
            existing_product = await db.products.find_one({"id": product.product_id})
            if not existing_product:
                continue
                
            package_product = {
                "id": str(uuid.uuid4()),
                "package_id": package_id,
                "product_id": product.product_id,
                "quantity": product.quantity,
                "created_at": datetime.now(timezone.utc)
            }
            package_products.append(package_product)
        
        if package_products:
            await db.package_products.insert_many(package_products)
        
        return {"success": True, "message": f"{len(package_products)} ürün pakete eklendi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding products to package: {e}")
        raise HTTPException(status_code=500, detail="Ürünler pakete eklenemedi")
    try:
        products = await db.products.find({"is_favorite": True}).sort("name", 1).to_list(None)
        return [Product(**product) for product in products]
    except Exception as e:
        logger.error(f"Error getting favorite products: {e}")
        raise HTTPException(status_code=500, detail="Favori ürünler getirilemedi")

@api_router.post("/companies/{company_id}/upload-excel")  
async def upload_excel(company_id: str, file: UploadFile = File(...), currency: str = Form(None), discount: str = Form("0")):
    """Upload Excel file for a company with smart update system"""
    try:
        # Verify company exists
        company = await db.companies.find_one({"id": company_id})
        if not company:
            raise HTTPException(status_code=404, detail="Firma bulunamadı")
        
        # Check file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Sadece Excel dosyaları (.xlsx, .xls) kabul edilir")
        
        # Read file content
        file_content = await file.read()
        
        # Try color-based parsing first, then fall back to traditional parsing
        try:
            # Color-based parsing
            products_data = ColorBasedExcelService.parse_colored_excel(file_content, company['name'])
            logger.info(f"Color-based parsing successful: {len(products_data)} products")
        except Exception as color_parse_error:
            logger.warning(f"Color-based parsing failed: {color_parse_error}")
            # Fall back to traditional parsing
            products_data = excel_service.parse_excel_file(file_content)
            logger.info(f"Traditional parsing used: {len(products_data)} products")
        
        if not products_data:
            raise HTTPException(status_code=400, detail="Excel dosyasında geçerli ürün verisi bulunamadı")
        
        # Handle user-selected currency override
        user_selected_currency = None
        if currency and currency.upper() in ['USD', 'EUR', 'TRY']:
            user_selected_currency = currency.upper()
            logger.info(f"User selected currency override: {user_selected_currency}")
        
        # Handle discount percentage
        discount_percentage = 0.0
        try:
            if discount and discount.strip():
                discount_percentage = float(discount)
                if discount_percentage < 0 or discount_percentage > 100:
                    raise ValueError("Discount must be between 0 and 100")
                logger.info(f"User selected discount: {discount_percentage}%")
        except ValueError as e:
            logger.error(f"Invalid discount value: {discount}, error: {e}")
            raise HTTPException(status_code=400, detail=f"Geçersiz iskonto değeri: {discount}")
        
        # Get current exchange rates
        await currency_service.get_exchange_rates()
        
        # Initialize counters and tracking
        new_products = 0
        updated_products = 0
        price_changes = []
        currency_distribution = {}
        created_products = []
        
        # Get existing products for this company for comparison
        existing_products_cursor = db.products.find({"company_id": company_id})
        existing_products = {product['name']: product async for product in existing_products_cursor}
        
        # Process and save products with smart update
        for product_data in products_data:
            try:
                # Handle company management for color-based parsing
                target_company_id = company_id
                target_company_name = company['name']
                
                # If product has a different company name (from color-based parsing)
                if (product_data.get('company_name') and 
                    product_data['company_name'] != company['name'] and
                    product_data['company_name'] != "Unknown"):
                    
                    # Check if this company already exists
                    existing_company = await db.companies.find_one({"name": product_data['company_name']})
                    if existing_company:
                        target_company_id = existing_company['id']
                        target_company_name = existing_company['name']
                    else:
                        # Create new company
                        new_company_dict = {
                            "id": str(uuid.uuid4()),
                            "name": product_data['company_name'],
                            "created_at": datetime.now(timezone.utc)
                        }
                        await db.companies.insert_one(new_company_dict)
                        target_company_id = new_company_dict['id']
                        target_company_name = new_company_dict['name']
                        logger.info(f"Created new company: {product_data['company_name']}")
                
                # Use user-selected currency if provided, otherwise use detected currency
                final_currency = user_selected_currency if user_selected_currency else product_data.get('currency', 'USD')
                
                # Apply discount if specified
                original_list_price = Decimal(str(product_data['list_price']))
                list_price = original_list_price  # Liste fiyatı orijinal fiyat olarak kalır
                
                # Calculate discounted price based on user discount percentage
                discounted_price = None
                if discount_percentage > 0:
                    # İskonto yüzdesi varsa, orijinal fiyattan indirim yap
                    discount_amount = original_list_price * (Decimal(str(discount_percentage)) / Decimal('100'))
                    discounted_price = original_list_price - discount_amount
                    logger.info(f"Applied {discount_percentage}% discount: {original_list_price} -> {discounted_price}")
                elif product_data.get('discounted_price'):
                    # Excel'de zaten indirimli fiyat varsa onu kullan
                    discounted_price = Decimal(str(product_data['discounted_price']))
                
                # Convert prices to TRY
                list_price_try = await currency_service.convert_to_try(list_price, final_currency)
                
                discounted_price_try = None
                if discounted_price:
                    discounted_price_try = await currency_service.convert_to_try(discounted_price, final_currency)
                
                # Count currency distribution (use final currency)
                currency = final_currency
                currency_distribution[currency] = currency_distribution.get(currency, 0) + 1
                
                # Check if product already exists (by name and company)
                product_name = product_data['name']
                if product_name in existing_products:
                    # Product exists - update it
                    existing_product = existing_products[product_name]
                    old_list_price = float(existing_product.get('list_price', 0))
                    new_list_price = float(product_data['list_price'])
                    
                    # Calculate price change
                    new_list_price = float(list_price)
                    if old_list_price != new_list_price:
                        price_change_amount = new_list_price - old_list_price
                        price_change_percent = ((new_list_price - old_list_price) / old_list_price * 100) if old_list_price > 0 else 0
                        
                        price_changes.append({
                            "product_name": product_name,
                            "old_price": old_list_price,
                            "new_price": new_list_price,
                            "change_amount": price_change_amount,
                            "change_percent": round(price_change_percent, 2),
                            "currency": currency,
                            "change_type": "increase" if price_change_amount > 0 else "decrease"
                        })
                    
                    # Update existing product
                    update_data = {
                        "brand": product_data.get('brand', ''),  # Marka güncellemesi
                        "list_price": float(list_price),
                        "discounted_price": float(discounted_price) if discounted_price else None,
                        "currency": final_currency,
                        "list_price_try": float(list_price_try),
                        "discounted_price_try": float(discounted_price_try) if discounted_price_try else None,
                        "updated_at": datetime.now(timezone.utc)
                    }
                    
                    await db.products.update_one(
                        {"id": existing_product['id']},
                        {"$set": update_data}
                    )
                    updated_products += 1
                    
                else:
                    # New product - create it
                    product_dict = {
                        "id": str(uuid.uuid4()),
                        "name": product_data['name'],
                        "company_id": target_company_id,
                        "brand": product_data.get('brand', ''),  # Marka alanı
                        "description": product_data.get('description'),
                        "image_url": None,
                        "list_price": float(list_price),
                        "discounted_price": float(discounted_price) if discounted_price else None,
                        "currency": final_currency,
                        "list_price_try": float(list_price_try),
                        "discounted_price_try": float(discounted_price_try) if discounted_price_try else None,
                        "created_at": datetime.now(timezone.utc)
                    }
                    
                    await db.products.insert_one(product_dict)
                    created_products.append(Product(**product_dict))
                    new_products += 1
                
            except Exception as e:
                logger.warning(f"Error processing product {product_data.get('name', 'Unknown')}: {e}")
                continue
        
        # Create upload history record
        upload_history = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "company_name": company['name'],
            "filename": file.filename,
            "upload_date": datetime.now(timezone.utc),
            "total_products": len(products_data),
            "new_products": new_products,
            "updated_products": updated_products,
            "currency_distribution": currency_distribution,
            "price_changes": price_changes,
            "status": "completed"
        }
        
        await db.upload_history.insert_one(upload_history)
        
        # Create detailed response message
        messages = []
        if new_products > 0:
            messages.append(f"{new_products} yeni ürün eklendi")
        if updated_products > 0:
            messages.append(f"{updated_products} ürün güncellendi")
        if price_changes:
            price_increases = len([c for c in price_changes if c['change_type'] == 'increase'])
            price_decreases = len([c for c in price_changes if c['change_type'] == 'decrease'])
            if price_increases > 0:
                messages.append(f"{price_increases} ürünün fiyatı zamlandı")
            if price_decreases > 0:
                messages.append(f"{price_decreases} ürünün fiyatı ucuzladı")
        
        message = ". ".join(messages) if messages else "Liste başarıyla yüklendi"
        
        return {
            "success": True,
            "message": message,
            "upload_id": upload_history["id"],
            "summary": {
                "total_products": len(products_data),
                "new_products": new_products,
                "updated_products": updated_products,
                "price_changes": len(price_changes),
                "currency_distribution": currency_distribution
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading Excel file: {e}")
        raise HTTPException(status_code=500, detail=f"Excel dosyası yüklenemedi: {str(e)}")

@api_router.get("/products/count")
async def get_products_count(
    company_id: Optional[str] = None,
    category_id: Optional[str] = None,
    search: Optional[str] = None
):
    """Get total count of products with optional filters"""
    try:
        query = {}
        if company_id:
            query["company_id"] = company_id
        if category_id:
            query["category_id"] = category_id
        if search:
            # Case-insensitive search in product name and description
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"description": {"$regex": search, "$options": "i"}}
            ]
            
        count = await db.products.count_documents(query)
        return {"count": count}
    except Exception as e:
        logger.error(f"Error getting products count: {e}")
        raise HTTPException(status_code=500, detail="Ürün sayısı getirilemedi")

@api_router.get("/products", response_model=List[Product])
async def get_products(
    company_id: Optional[str] = None,
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 100,
    skip_pagination: bool = False  # For backward compatibility
):
    """Get products with optimized pagination, filtering by company, category, or search term"""
    try:
        query = {}
        if company_id:
            query["company_id"] = company_id
        if category_id:
            query["category_id"] = category_id
        if search:
            # Enhanced text search with index utilization
            search_term = search.strip()
            if len(search_term) >= 2:  # Only search for 2+ characters for performance
                query["$text"] = {"$search": search_term}
            else:
                # Fallback for short searches
                query["$or"] = [
                    {"name": {"$regex": f"^{search}", "$options": "i"}},
                    {"brand": {"$regex": f"^{search}", "$options": "i"}}
                ]
        
        # Optimize sorting: favorites first, then by name with index utilization
        sort_criteria = [("is_favorite", -1), ("name", 1)]
        
        # For backward compatibility - if skip_pagination is true, return all
        if skip_pagination:
            # For large datasets, still apply reasonable limits
            max_limit = 5000
            products = await db.products.find(query).sort(sort_criteria).limit(max_limit).to_list(max_limit)
        else:
            # Enhanced pagination with performance optimizations
            skip = (page - 1) * limit
            
            # Use hint for index optimization when applicable
            cursor = db.products.find(query).sort(sort_criteria).skip(skip).limit(limit)
            
            # Add index hint for better performance on large datasets
            if not search:  # Only use hint when not doing text search
                cursor = cursor.hint([("is_favorite", -1), ("name", 1)])
            
            products = await cursor.to_list(limit)
            
        return [Product(**product) for product in products]
    except Exception as e:
        logger.error(f"Error getting products: {e}")
        # Fallback to basic query if optimization fails
        try:
            basic_query = {}
            if company_id:
                basic_query["company_id"] = company_id
            if category_id:
                basic_query["category_id"] = category_id
            
            skip = (page - 1) * limit if not skip_pagination else 0
            products = await db.products.find(basic_query).sort([("name", 1)]).skip(skip).limit(limit).to_list(limit)
            return [Product(**product) for product in products]
        except Exception as fallback_error:
            logger.error(f"Fallback query also failed: {fallback_error}")
            raise HTTPException(status_code=500, detail="Ürünler getirilemedi")

@api_router.post("/products", response_model=Product)
async def create_product(product: ProductCreate):
    """Create a new product manually"""
    try:
        # Verify company exists
        company = await db.companies.find_one({"id": product.company_id})
        if not company:
            raise HTTPException(status_code=404, detail="Firma bulunamadı")
        
        # Verify category exists (if provided)
        if product.category_id:
            category = await db.categories.find_one({"id": product.category_id})
            if not category:
                raise HTTPException(status_code=404, detail="Kategori bulunamadı")
        
        # Get current exchange rates for TRY conversion
        await currency_service.get_exchange_rates()
        
        # Convert prices to TRY
        try:
            list_price_try = await currency_service.convert_to_try(
                product.list_price, 
                product.currency
            )
        except Exception as e:
            logger.warning(f"Failed to convert price to TRY for product {product.name}: {e}")
            # Fallback: Use original price as TRY if conversion fails
            list_price_try = product.list_price
        
        discounted_price_try = None
        if product.discounted_price:
            try:
                discounted_price_try = await currency_service.convert_to_try(
                    product.discounted_price, 
                    product.currency
                )
            except Exception as e:
                logger.warning(f"Failed to convert discounted price to TRY for product {product.name}: {e}")
                discounted_price_try = product.discounted_price
        
        # Create product
        product_dict = {
            "id": str(uuid.uuid4()),
            "name": product.name,
            "company_id": product.company_id,
            "category_id": product.category_id,
            "description": product.description,
            "image_url": product.image_url,
            "list_price": float(product.list_price),
            "discounted_price": float(product.discounted_price) if product.discounted_price else None,
            "currency": product.currency.upper(),
            "list_price_try": float(list_price_try),
            "discounted_price_try": float(discounted_price_try) if discounted_price_try else None,
            "is_favorite": product.is_favorite,  # Include is_favorite field
            "created_at": datetime.now(timezone.utc)
        }
        
        await db.products.insert_one(product_dict)
        
        return Product(**product_dict)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating product: {e}")
        raise HTTPException(status_code=500, detail="Ürün oluşturulamadı")



@api_router.post("/refresh-prices")
async def refresh_prices():
    """Refresh all product prices with current exchange rates"""
    try:
        # Get fresh exchange rates
        await currency_service.get_exchange_rates()
        
        # Get all products
        products = await db.products.find().to_list(None)
        updated_count = 0
        
        for product in products:
            try:
                if product['currency'] != 'TRY':
                    # Convert prices to TRY
                    list_price_try = await currency_service.convert_to_try(
                        Decimal(str(product['list_price'])), 
                        product['currency']
                    )
                    
                    discounted_price_try = None
                    if product.get('discounted_price'):
                        discounted_price_try = await currency_service.convert_to_try(
                            Decimal(str(product['discounted_price'])), 
                            product['currency']
                        )
                    
                    # Update product
                    await db.products.update_one(
                        {"id": product["id"]},
                        {
                            "$set": {
                                "list_price_try": float(list_price_try),
                                "discounted_price_try": float(discounted_price_try) if discounted_price_try else None
                            }
                        }
                    )
                    updated_count += 1
                    
            except Exception as e:
                logger.warning(f"Error updating product {product.get('name', 'Unknown')}: {e}")
                continue
        
        return {
            "success": True,
            "message": f"{updated_count} ürünün fiyatı güncellendi",
            "updated_count": updated_count
        }
        
    except Exception as e:
        logger.error(f"Error refreshing prices: {e}")
        raise HTTPException(status_code=500, detail="Fiyatlar güncellenemedi")

# Upload History Endpoints
@api_router.get("/companies/{company_id}/upload-history", response_model=List[UploadHistoryResponse])
async def get_company_upload_history(company_id: str):
    """Get upload history for a specific company"""
    try:
        # Verify company exists
        company = await db.companies.find_one({"id": company_id})
        if not company:
            raise HTTPException(status_code=404, detail="Firma bulunamadı")
        
        # Get upload history for this company, sorted by date (newest first)
        upload_history = await db.upload_history.find(
            {"company_id": company_id}
        ).sort("upload_date", -1).to_list(None)
        
        return [UploadHistoryResponse(**history) for history in upload_history]
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting upload history: {e}")
        raise HTTPException(status_code=500, detail="Upload geçmişi getirilemedi")

@api_router.get("/upload-history/{upload_id}", response_model=UploadHistoryResponse)
async def get_upload_details(upload_id: str):
    """Get detailed information about a specific upload"""
    try:
        upload = await db.upload_history.find_one({"id": upload_id})
        if not upload:
            raise HTTPException(status_code=404, detail="Upload bulunamadı")
        
        return UploadHistoryResponse(**upload)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting upload details: {e}")
        raise HTTPException(status_code=500, detail="Upload detayları getirilemedi")

@api_router.get("/upload-history", response_model=List[UploadHistoryResponse])
async def get_all_upload_history():
    """Get all upload history across all companies"""
    try:
        upload_history = await db.upload_history.find().sort("upload_date", -1).to_list(None)
        return [UploadHistoryResponse(**history) for history in upload_history]
        
    except Exception as e:
        logger.error(f"Error getting all upload history: {e}")
        raise HTTPException(status_code=500, detail="Upload geçmişi getirilemedi")

@api_router.post("/upload-history/{upload_id}/change-currency")
async def change_upload_currency(upload_id: str, new_currency: str):
    """Change currency for all products in a specific upload"""
    try:
        # Validate currency
        valid_currencies = ['USD', 'EUR', 'TRY', 'GBP']
        if new_currency not in valid_currencies:
            raise HTTPException(status_code=400, detail=f"Geçersiz para birimi. Geçerli seçenekler: {', '.join(valid_currencies)}")
        
        # Get upload history
        upload = await db.upload_history.find_one({"id": upload_id})
        if not upload:
            raise HTTPException(status_code=404, detail="Upload bulunamadı")
        
        # Get current exchange rates
        await currency_service.get_exchange_rates()
        
        # Find all products uploaded in this batch
        # We'll identify products by upload date range (within 5 minutes of upload)
        upload_date = upload['upload_date']
        start_time = upload_date - timedelta(minutes=5)
        end_time = upload_date + timedelta(minutes=5)
        
        # Get products from this company within the time range
        company_id = upload['company_id']
        products_cursor = db.products.find({
            "company_id": company_id,
            "created_at": {"$gte": start_time, "$lte": end_time}
        })
        products = await products_cursor.to_list(None)
        
        if not products:
            # Alternative method: get products by estimated count
            # This is less precise but more reliable
            company_products_cursor = db.products.find({"company_id": company_id}).sort("created_at", -1)
            all_company_products = await company_products_cursor.to_list(None)
            
            # Take approximately the number of products that were in the upload
            expected_count = upload.get('total_products', 0)
            if expected_count > 0 and len(all_company_products) >= expected_count:
                products = all_company_products[:expected_count]
            else:
                raise HTTPException(status_code=404, detail="Bu upload'a ait ürünler bulunamadı")
        
        updated_count = 0
        price_changes = []
        
        # Update each product's currency (PRESERVE PRICE VALUES, ONLY CHANGE CURRENCY LABEL)
        for product in products:
            try:
                old_currency = product.get('currency', 'TRY')
                old_list_price = product.get('list_price', 0)
                old_discounted_price = product.get('discounted_price')
                
                # Skip if already in target currency
                if old_currency == new_currency:
                    continue
                
                ## IMPORTANT: Keep the same price values, only change currency label
                ## This is for cases where Excel had correct prices but wrong currency was detected
                
                # Keep the same numeric values, just change the currency
                new_list_price = old_list_price  # Same value!
                new_discounted_price = old_discounted_price  # Same value!
                
                # Recalculate TRY prices based on new currency (for internal calculations)
                new_list_price_try = await currency_service.convert_to_try(
                    Decimal(str(new_list_price)), new_currency
                )
                
                new_discounted_price_try = None
                if new_discounted_price:
                    new_discounted_price_try = await currency_service.convert_to_try(
                        Decimal(str(new_discounted_price)), new_currency
                    )
                
                # Update product in database
                update_data = {
                    "currency": new_currency,
                    "list_price": float(new_list_price),  # Same numeric value
                    "list_price_try": float(new_list_price_try),  # Recalculated for TRY
                    "updated_at": datetime.now(timezone.utc)
                }
                
                if new_discounted_price:
                    update_data["discounted_price"] = float(new_discounted_price)  # Same numeric value
                    update_data["discounted_price_try"] = float(new_discounted_price_try)  # Recalculated for TRY
                    
                    await db.products.update_one(
                        {"id": product['id']},
                        {"$set": update_data}
                    )
                    
                    updated_count += 1
                    
                    # Track currency change (prices stay the same, only currency label changes)
                    price_changes.append({
                        "product_name": product['name'],
                        "old_currency": old_currency,
                        "new_currency": new_currency,
                        "price_value": float(old_list_price),  # Same value in both currencies
                        "change_type": "currency_label_only"
                    })
                    
            except Exception as e:
                logger.warning(f"Error updating product {product.get('name', 'Unknown')}: {e}")
                continue
        
        # Update upload history to reflect the currency change
        await db.upload_history.update_one(
            {"id": upload_id},
            {
                "$set": {
                    "currency_changes": price_changes,
                    "last_currency_update": datetime.now(timezone.utc)
                }
            }
        )
        
        return {
            "success": True,
            "message": f"{updated_count} ürünün para birimi {new_currency} olarak güncellendi (fiyat değerleri aynı kaldı)",
            "updated_count": updated_count,
            "currency_changes": price_changes[:10]  # Show first 10 changes
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error changing upload currency: {e}")
        raise HTTPException(status_code=500, detail=f"Para birimi güncellenemedi: {str(e)}")

# Category Groups CRUD Operations
@api_router.get("/category-groups")
async def get_category_groups():
    """Get all category groups sorted by sort_order, then by name"""
    try:
        groups = []
        # Kategorileri önce sort_order'a, sonra name'e göre sırala
        async for group in db.category_groups.find().sort([("sort_order", 1), ("name", 1)]):
            # Remove MongoDB _id field
            group.pop('_id', None)
            groups.append(group)
        return groups
    except Exception as e:
        logger.error(f"Error fetching category groups: {e}")
        raise HTTPException(status_code=500, detail="Kategori grupları getirilemedi")

@api_router.post("/category-groups")
async def create_category_group(group_data: CategoryGroupCreate):
    """Create a new category group"""
    try:
        group = {
            "id": str(uuid.uuid4()),
            "name": group_data.name,
            "description": group_data.description,
            "color": group_data.color or "#6B7280",
            "category_ids": group_data.category_ids,
            "created_at": datetime.now(timezone.utc)
        }
        
        result = await db.category_groups.insert_one(group)
        # Remove MongoDB _id field for response
        group.pop('_id', None)
        return {"success": True, "message": "Kategori grubu oluşturuldu", "group": group}
    except Exception as e:
        logger.error(f"Error creating category group: {e}")
        raise HTTPException(status_code=500, detail="Kategori grubu oluşturulamadı")

@api_router.put("/category-groups/{group_id}")
async def update_category_group(group_id: str, group_data: CategoryGroupUpdate):
    """Update a category group"""
    try:
        update_dict = {}
        if group_data.name is not None:
            update_dict["name"] = group_data.name
        if group_data.description is not None:
            update_dict["description"] = group_data.description
        if group_data.color is not None:
            update_dict["color"] = group_data.color
        if group_data.category_ids is not None:
            update_dict["category_ids"] = group_data.category_ids
        if group_data.sort_order is not None:
            update_dict["sort_order"] = group_data.sort_order
            
        if not update_dict:
            raise HTTPException(status_code=400, detail="Güncellenecek alan belirtilmedi")
            
        result = await db.category_groups.update_one(
            {"id": group_id},
            {"$set": update_dict}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Kategori grubu bulunamadı")
            
        return {"success": True, "message": "Kategori grubu güncellendi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating category group: {e}")
        raise HTTPException(status_code=500, detail="Kategori grubu güncellenemedi")

@api_router.delete("/category-groups/{group_id}")
async def delete_category_group(group_id: str):
    """Delete a category group"""
    try:
        result = await db.category_groups.delete_one({"id": group_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Kategori grubu bulunamadı")
            
        return {"success": True, "message": "Kategori grubu silindi"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting category group: {e}")
        raise HTTPException(status_code=500, detail="Kategori grubu silinemedi")

@api_router.post("/category-groups/reorder")
async def reorder_category_groups(group_orders: List[Dict[str, Any]]):
    """Kategori gruplarının sırasını toplu güncelle"""
    try:
        # group_orders: [{"id": "group1", "sort_order": 1}, {"id": "group2", "sort_order": 2}, ...]
        for item in group_orders:
            group_id = item.get("id")
            sort_order = item.get("sort_order", 0)
            
            if group_id:
                await db.category_groups.update_one(
                    {"id": group_id},
                    {"$set": {"sort_order": sort_order}}
                )
        
        # Return updated category groups sorted by new order
        groups = []
        async for group in db.category_groups.find().sort([("sort_order", 1), ("name", 1)]):
            group.pop('_id', None)
            groups.append(group)
        
        return {
            "success": True,
            "message": "Kategori grubu sıralaması güncellendi",
            "category_groups": groups
        }
        
    except Exception as e:
        logger.error(f"Error reordering category groups: {e}")
        raise HTTPException(status_code=500, detail="Kategori grubu sıralaması güncellenemedi")

# Static file serving for MongoDB Atlas migration files
@api_router.get("/atlas-downloads")
async def downloads_page():
    """Serve the downloads index page"""
    try:
        with open("/app/downloads/index.html", "r", encoding="utf-8") as f:
            content = f.read()
        return StreamingResponse(
            io.StringIO(content),
            media_type="text/html; charset=utf-8"
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Downloads page not found")

@api_router.get("/atlas-downloads/{filename}")
async def download_file(filename: str):
    """Serve JSON files for download"""
    file_path = f"/app/downloads/{filename}"
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"File {filename} not found")
    
    if not filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="Only JSON files are allowed for download")
    
    return FileResponse(
        file_path,
        media_type="application/json",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Mount static files directory for any other files
try:
    app.mount("/downloads", StaticFiles(directory="/app/downloads"), name="downloads")
except Exception as e:
    logger.warning(f"Could not mount static files: {e}")

# Authentication Endpoints
@api_router.post("/auth/login", response_model=LoginResponse)
async def login(login_request: LoginRequest, response: JSONResponse):
    """User login endpoint"""
    try:
        # Find user in database
        user = await db.users.find_one({"username": login_request.username})
        if not user:
            raise HTTPException(status_code=401, detail="Geçersiz kullanıcı adı veya şifre")
        
        # Verify password
        if not auth_service.verify_password(login_request.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Geçersiz kullanıcı adı veya şifre")
        
        # Check if user is active
        if not user.get("is_active", True):
            raise HTTPException(status_code=401, detail="Hesap devre dışı")
        
        # Create session
        session_token = auth_service.create_session(login_request.username)
        
        # Set session cookie
        response = JSONResponse(
            content={
                "success": True,
                "message": "Başarıyla giriş yapıldı",
                "session_token": session_token
            }
        )
        response.set_cookie(
            key="session_token",
            value=session_token,
            max_age=86400,  # 24 hours in seconds
            httponly=True,
            secure=False,  # Set True in production with HTTPS
            samesite="lax"
        )
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {e}")
        raise HTTPException(status_code=500, detail="Giriş işlemi sırasında hata oluştu")

@api_router.post("/auth/logout")
async def logout(session_token: Optional[str] = Cookie(None)):
    """User logout endpoint"""
    try:
        if session_token:
            auth_service.logout(session_token)
        
        response = JSONResponse(content={"success": True, "message": "Başarıyla çıkış yapıldı"})
        response.delete_cookie("session_token")
        return response
        
    except Exception as e:
        logger.error(f"Logout error: {e}")
        response = JSONResponse(content={"success": True, "message": "Çıkış yapıldı"})
        response.delete_cookie("session_token")
        return response

@api_router.get("/auth/check")
async def check_auth(current_user: str = Depends(get_current_user_optional)):
    """Check authentication status"""
    return {
        "authenticated": current_user is not None,
        "username": current_user
    }

# Include the router in the main app
app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)