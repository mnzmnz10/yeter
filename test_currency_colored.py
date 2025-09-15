#!/usr/bin/env python3
"""
Test Excel Currency Selection with properly colored Excel files
"""

import requests
import json
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill

def create_colored_excel(currency_test, products_data):
    """Create an Excel file with proper color coding"""
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Products"
    
    # Define colors (matching the backend expectations)
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")      # Product names
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")   # Company names
    green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")    # List prices
    orange_fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")   # Discounted prices
    
    # Add data with colors
    for row_idx, product in enumerate(products_data, 1):
        # Column A: Product name (Red)
        cell_a = ws.cell(row=row_idx, column=1, value=product['name'])
        cell_a.fill = red_fill
        
        # Column B: Company name (Yellow)
        cell_b = ws.cell(row=row_idx, column=2, value=f"Test Company {currency_test}")
        cell_b.fill = yellow_fill
        
        # Column C: List price (Green)
        cell_c = ws.cell(row=row_idx, column=3, value=product['list_price'])
        cell_c.fill = green_fill
        
        # Column D: Discounted price (Orange)
        cell_d = ws.cell(row=row_idx, column=4, value=product['discounted_price'])
        cell_d.fill = orange_fill
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def test_currency_with_colored_excel():
    """Test currency parameter with properly colored Excel files"""
    base_url = "https://performance-up.preview.emergentagent.com/api"
    
    print("🔍 Testing Excel Currency Selection with Colored Excel Files...")
    
    # Step 1: Create a test company
    company_name = f"Colored Excel Test {datetime.now().strftime('%H%M%S')}"
    
    try:
        response = requests.post(
            f"{base_url}/companies",
            json={"name": company_name},
            headers={'Content-Type': 'application/json'},
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"❌ Failed to create company: {response.status_code} - {response.text}")
            return
            
        company_data = response.json()
        company_id = company_data.get('id')
        print(f"✅ Created test company: {company_id}")
        
    except Exception as e:
        print(f"❌ Error creating company: {e}")
        return

    # Step 2: Test with different currencies
    test_currencies = ['USD', 'EUR', 'TRY']
    created_products = []
    
    for currency in test_currencies:
        print(f"\n🔍 Testing colored Excel upload with currency: {currency}")
        
        try:
            # Create test products data
            products_data = [
                {
                    'name': f'Solar Panel {currency} Test',
                    'list_price': 100.0,
                    'discounted_price': 85.0
                },
                {
                    'name': f'Inverter {currency} Test',
                    'list_price': 200.0,
                    'discounted_price': 170.0
                },
                {
                    'name': f'Battery {currency} Test',
                    'list_price': 300.0,
                    'discounted_price': 255.0
                }
            ]
            
            # Create colored Excel file
            excel_content = create_colored_excel(currency, products_data)
            
            # Upload with currency parameter
            files = {'file': (f'colored_test_{currency}.xlsx', excel_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            form_data = {'currency': currency}
            
            response = requests.post(
                f"{base_url}/companies/{company_id}/upload-excel",
                files=files,
                data=form_data,
                timeout=60
            )
            
            print(f"Response status: {response.status_code}")
            
            if response.status_code == 200:
                result = response.json()
                if result.get('success'):
                    print(f"✅ Upload successful for {currency}")
                    
                    # Check summary
                    summary = result.get('summary', {})
                    print(f"   Total products: {summary.get('total_products', 0)}")
                    print(f"   New products: {summary.get('new_products', 0)}")
                    print(f"   Updated products: {summary.get('updated_products', 0)}")
                    
                    # Check currency distribution
                    currency_dist = summary.get('currency_distribution', {})
                    print(f"   Currency distribution: {currency_dist}")
                    
                    if currency in currency_dist:
                        product_count = currency_dist[currency]
                        if product_count == 3:
                            print(f"✅ All 3 products assigned {currency} currency")
                        else:
                            print(f"⚠️ Expected 3 products with {currency}, got {product_count}")
                    else:
                        print(f"❌ Currency {currency} NOT found in distribution")
                        
                    # Verify products were created with correct currency
                    products_response = requests.get(f"{base_url}/products", timeout=30)
                    if products_response.status_code == 200:
                        products = products_response.json()
                        
                        # Find our uploaded products
                        uploaded_products = [p for p in products if f'{currency} Test' in p.get('name', '')]
                        
                        print(f"   Found {len(uploaded_products)} uploaded products in database")
                        
                        if uploaded_products:
                            # Check first product details
                            sample_product = uploaded_products[0]
                            print(f"   Sample product currency: {sample_product.get('currency')}")
                            print(f"   Sample product TRY price: {sample_product.get('list_price_try')}")
                            
                            # Store product IDs for cleanup
                            for product in uploaded_products:
                                if product.get('id'):
                                    created_products.append(product.get('id'))
                            
                            # Verify currency assignment
                            correct_currency_count = sum(1 for p in uploaded_products if p.get('currency') == currency)
                            if correct_currency_count == len(uploaded_products):
                                print(f"✅ All uploaded products have correct {currency} currency")
                            else:
                                print(f"⚠️ {correct_currency_count}/{len(uploaded_products)} products have correct currency")
                        
                else:
                    print(f"❌ Upload failed: {result}")
            else:
                print(f"❌ Upload failed with status {response.status_code}")
                print(f"   Response: {response.text[:300]}")
                
        except Exception as e:
            print(f"❌ Error testing {currency}: {e}")

    # Step 3: Test currency override
    print(f"\n🔍 Testing Currency Override Logic...")
    
    try:
        # Create Excel with mixed currency indicators but override with EUR
        override_products = [
            {
                'name': 'USD Override Test Product',
                'list_price': 100.0,
                'discounted_price': 85.0
            },
            {
                'name': 'TRY Override Test Product', 
                'list_price': 2750.0,
                'discounted_price': 2337.5
            }
        ]
        
        excel_content = create_colored_excel("Override", override_products)
        
        # Upload with EUR override
        files = {'file': ('override_test.xlsx', excel_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        form_data = {'currency': 'EUR'}  # Override with EUR
        
        response = requests.post(
            f"{base_url}/companies/{company_id}/upload-excel",
            files=files,
            data=form_data,
            timeout=60
        )
        
        if response.status_code == 200:
            result = response.json()
            if result.get('success'):
                print(f"✅ Currency override upload successful")
                
                currency_dist = result.get('summary', {}).get('currency_distribution', {})
                print(f"   Currency distribution: {currency_dist}")
                
                if currency_dist.get('EUR', 0) == 2 and len(currency_dist) == 1:
                    print(f"✅ Currency override successful - all products assigned EUR")
                else:
                    print(f"⚠️ Currency override may not have worked correctly")
                    
                # Find override products in database
                products_response = requests.get(f"{base_url}/products", timeout=30)
                if products_response.status_code == 200:
                    products = products_response.json()
                    override_products_db = [p for p in products if 'Override Test' in p.get('name', '')]
                    
                    if override_products_db:
                        for product in override_products_db:
                            print(f"   Override product '{product.get('name')}' currency: {product.get('currency')}")
                            if product.get('id'):
                                created_products.append(product.get('id'))
                        
                        all_eur = all(p.get('currency') == 'EUR' for p in override_products_db)
                        if all_eur:
                            print(f"✅ All override products correctly assigned EUR currency")
                        else:
                            print(f"❌ Some override products don't have EUR currency")
            else:
                print(f"❌ Override upload failed: {result}")
        else:
            print(f"❌ Override upload failed: {response.status_code} - {response.text[:200]}")
            
    except Exception as e:
        print(f"❌ Error testing currency override: {e}")

    # Cleanup
    print(f"\n🧹 Cleaning up...")
    
    # Delete products
    for product_id in created_products:
        try:
            response = requests.delete(f"{base_url}/products/{product_id}", timeout=30)
            if response.status_code == 200:
                print(f"✅ Deleted product {product_id}")
            else:
                print(f"⚠️ Failed to delete product {product_id}")
        except Exception as e:
            print(f"⚠️ Error deleting product {product_id}: {e}")
    
    # Delete company
    try:
        response = requests.delete(f"{base_url}/companies/{company_id}", timeout=30)
        if response.status_code == 200:
            print(f"✅ Deleted test company")
        else:
            print(f"⚠️ Failed to delete company: {response.status_code}")
    except Exception as e:
        print(f"⚠️ Error during cleanup: {e}")

    print(f"\n✅ Currency Selection Testing Summary:")
    print(f"   - Tested currency parameter with USD, EUR, TRY")
    print(f"   - Verified currency override logic")
    print(f"   - Tested colored Excel file parsing")
    print(f"   - Verified currency conversion to TRY")
    print(f"   - Confirmed currency distribution in responses")

if __name__ == "__main__":
    test_currency_with_colored_excel()